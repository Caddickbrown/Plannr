import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import pymssql
from dotenv import load_dotenv
import sys
from sqlalchemy import create_engine, text
import psutil
import gc
from io import BytesIO

# Load environment variables
load_dotenv('db_credentials.env')

# Performance tracking utilities
class PerformanceTracker:
    """Track performance metrics across different phases"""
    
    def __init__(self):
        self.phases = {}
        self.current_phase = None
        self.phase_start_time = None
        self.memory_usage = []
        self.process = psutil.Process()
    
    def start_phase(self, phase_name):
        """Start timing a new phase"""
        if self.current_phase:
            self.end_phase()
        
        self.current_phase = phase_name
        self.phase_start_time = time.time()
        memory_info = self.process.memory_info()
        self.memory_usage.append({
            'phase': phase_name,
            'memory_mb': memory_info.rss / 1024 / 1024,
            'timestamp': time.time()
        })
    
    def end_phase(self):
        """End the current phase and record timing"""
        if self.current_phase and self.phase_start_time:
            duration = time.time() - self.phase_start_time
            if self.current_phase not in self.phases:
                self.phases[self.current_phase] = []
            self.phases[self.current_phase].append(duration)
            
            # Record final memory usage for this phase
            memory_info = self.process.memory_info()
            self.memory_usage.append({
                'phase': f"{self.current_phase}_end",
                'memory_mb': memory_info.rss / 1024 / 1024,
                'timestamp': time.time()
            })
            
            self.current_phase = None
            self.phase_start_time = None
    
    def get_phase_summary(self):
        """Get summary of all phases"""
        summary = {}
        for phase, times in self.phases.items():
            if times:
                summary[phase] = {
                    'total_time': sum(times),
                    'avg_time': sum(times) / len(times),
                    'count': len(times),
                    'min_time': min(times),
                    'max_time': max(times)
                }
        return summary
    
    def get_memory_summary(self):
        """Get memory usage summary"""
        if not self.memory_usage:
            return {}
        
        memory_values = [m['memory_mb'] for m in self.memory_usage]
        return {
            'peak_memory_mb': max(memory_values),
            'avg_memory_mb': sum(memory_values) / len(memory_values),
            'initial_memory_mb': self.memory_usage[0]['memory_mb'] if self.memory_usage else 0,
            'final_memory_mb': self.memory_usage[-1]['memory_mb'] if self.memory_usage else 0
        }
    
    def cleanup(self):
        """Clean up and end any current phase"""
        if self.current_phase:
            self.end_phase()

# Global performance tracker
performance_tracker = PerformanceTracker()

def timing_decorator(phase_name):
    """Decorator to automatically time function execution"""
    def decorator(func):
        def wrapper(*args, **kwargs):
            performance_tracker.start_phase(phase_name)
            try:
                result = func(*args, **kwargs)
                return result
            finally:
                performance_tracker.end_phase()
        return wrapper
    return decorator

# VERSION INFO
VERSION = "v2.0.0"
VERSION_DATE = "2025-07-29"
DEBUG_MODE = False
DEBUG_COMPONENT_PART = None  # Set to a specific part number (as string) to track, e.g. "8034855"
DEBUG_SO_NUMBER = 9682591
      # Set to a specific SO number (as string) to track, e.g. "9678417"

# Global variables
quick_analysis_excel_buffer = None

# Database configuration
DB_HOST = os.getenv('DB_HOST', 'localhost')
DB_USER = os.getenv('DB_USER', 'myuser')
DB_PASSWORD = os.getenv('DB_PASSWORD', 'mypassword')
DB_NAME = os.getenv('DB_NAME', 'mydatabase')
DB_PORT = os.getenv('DB_PORT', '1433')

# Print debug configuration at startup
if DEBUG_MODE:
    print("\n" + "="*60)
    print("🔍 DEBUG MODE ACTIVATED")
    print("="*60)
    if DEBUG_COMPONENT_PART is None and DEBUG_SO_NUMBER is None:
        print("⚠️  NO DEBUG FILTERS SET")
        print("   Set DEBUG_COMPONENT_PART or DEBUG_SO_NUMBER to see detailed output")
        print("   Example: DEBUG_COMPONENT_PART = '8039831' or DEBUG_SO_NUMBER = '9678487'")
        print("   Current settings will show general processing info only")
    else:
        print("🎯 DEBUG FILTERS ACTIVE:")
        print(f"   Component Part: {DEBUG_COMPONENT_PART if DEBUG_COMPONENT_PART else 'Not specified'}")
        print(f"   SO Number: {DEBUG_SO_NUMBER if DEBUG_SO_NUMBER else 'Not specified'}")
    print("="*60)
    print("Debug output will appear in the terminal/console during processing")
    print("="*60 + "\n")
    
    # Show data source information
    print("📊 DATA SOURCES AND COLUMN MAPPINGS:")
    print("   📦 IPIS Table (Stock):")
    print("     - Column: 'PART_NO' → Used for stock lookup")
    print("     - Column: 'Available Qty' → Stock quantity")
    print("     - Logic: Grouped by PART_NO, sum of Available Qty")
    print("")
    print("   🔒 Component Demand Table (Committed):")
    print("     - Column: 'Component Part Number' → Component identifier")
    print("     - Column: 'Component Qty Required' → Committed quantity")
    print("     - Logic: Grouped by Component Part Number, sum of Component Qty Required")
    print("")
    print("   📋 Planned Demand Table (BOM):")
    print("     - Column: 'SO Number' → Shop Order identifier")
    print("     - Column: 'Component Part Number' → Component identifier")
    print("     - Column: 'Component Qty Required' → Required quantity")
    print("     - Logic: Filtered by SO Number to get BOM components")
    print("")
    print("   📄 POs Table (Purchase Orders):")
    print("     - Column: 'Part Number' → Part identifier")
    print("     - Column: 'Qty Due' → Quantity on order")
    print("     - Column: 'Promised Due Date' → Expected delivery")
    print("     - Logic: Filtered by Part Number and future dates")
    print("")
    print("   ⏱️ Hours Table (Labor Standards):")
    print("     - Column: 'PART_NO' → Part identifier")
    print("     - Column: 'Hours per Unit' → Labor hours")
    print("     - Logic: Grouped by PART_NO, sum of Hours per Unit")
    print("")
    print("   📋 Demand Table (Main Orders):")
    print("     - Column: 'SO No' → Shop Order identifier")
    print("     - Column: 'Part No' → Parent part identifier")
    print("     - Column: 'Rev Qty Due' → Order quantity")
    print("     - Column: 'Start Date' → Order start date")
    print("     - Column: 'Planner' → Planner code")
    print("="*60 + "\n")

def get_database_connection():
    """Create and return a SQLAlchemy engine for SQL Server"""
    try:
        # Create connection string for SQL Server
        connection_string = f"mssql+pyodbc://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}?driver=ODBC+Driver+17+for+SQL+Server"
        
        # Create SQLAlchemy engine with SSMS-like parameters
        engine = create_engine(
            connection_string,
            connect_args={
                'appname': 'Microsoft SQL Server Management Studio - Query',
                'charset': 'utf8',
                'login_timeout': 30,
                'timeout': 30
            },
            echo=False  # Set to True for SQL query logging
        )
        return engine
    except Exception as e:
        raise Exception(f"SQL Server connection failed: {str(e)}")

def execute_query(query, params=None):
    """Execute a SQL query and return results as a pandas DataFrame"""
    engine = None
    try:
        engine = get_database_connection()
        # Use SQLAlchemy text() for parameterized queries
        if params:
            df = pd.read_sql_query(text(query), engine, params=params)
        else:
            df = pd.read_sql_query(text(query), engine)
        return df
    except Exception as e:
        raise Exception(f"Query execution failed: {str(e)}")
    finally:
        if engine:
            engine.dispose()

@timing_decorator("Load Demand Data")
def load_demand_data():
    """Load demand data from database"""
    query = """
    SELECT
		so.ORDER_NO AS 'SO No',
		so.PART_NO AS 'Part No',
		ipt.PLANNER_BUYER AS 'Planner',
		ipt.PRIME_COMMODITY AS 'Comm Group',
		so.ROWSTATE AS 'Status',
		so.REVISED_START_DATE AS 'Start Date',
		so.REVISED_QTY_DUE AS 'Rev Qty Due',
		CASE WHEN RIGHT(so.PART_NO,1) = 'S' THEN 'Sterile' ELSE 'Non-Sterile' END AS 'Sterility'
    FROM IFS.SHOP_ORD_TAB AS so
	INNER JOIN IFS.INVENTORY_PART_TAB AS ipt
		ON so.CONTRACT = ipt.CONTRACT AND so.PART_NO = ipt.PART_NO
	WHERE so.CONTRACT = '2051'
		AND so.ROWSTATE IN ('Planned')
		AND ipt.PLANNER_BUYER IN ('3001','3801','5001','KIT SAMPLES','3802','3803','3804','3805')
	ORDER BY so.REVISED_START_DATE, so.ORDER_NO
	;
    """
    return execute_query(query)

@timing_decorator("Load Planned Demand Data")
def load_planned_demand_data():
    """Load planned demand (BOM) data from database"""
    query = """
    SELECT
		mac.ORDER_NO AS 'SO Number',
		so.PART_NO AS 'Kit Number',
		mac.PART_NO AS 'Component Part Number',
		mac.LINE_ITEM_NO AS 'Line Number',
		mac.DATE_REQUIRED AS 'Date Needed',
		mac.QTY_REQUIRED AS 'Component Qty Required'
	FROM IFS.SHOP_MATERIAL_ALLOC_TAB AS mac
	LEFT JOIN IFS.SHOP_ORD_TAB AS so
		ON mac.ORDER_NO = so.ORDER_NO AND mac.CONTRACT = so.CONTRACT
	WHERE
		so.CONTRACT = '2051'
		AND mac.ROWSTATE IN ('Planned')
	ORDER BY so.REVISED_START_DATE, mac.ORDER_NO
    """
    return execute_query(query)

@timing_decorator("Load Component Demand Data")
def load_component_demand_data():
    """Load committed component demand data from database"""
    query = """
    SELECT
		mac.ORDER_NO AS 'SO Number',
		so.PART_NO AS 'Kit Number',
		mac.PART_NO AS 'Component Part Number',
		mac.LINE_ITEM_NO AS 'Line Number',
		mac.DATE_REQUIRED AS 'Date Needed',
		mac.QTY_REQUIRED AS 'Component Qty Required'
	FROM IFS.SHOP_MATERIAL_ALLOC_TAB AS mac
	LEFT JOIN IFS.SHOP_ORD_TAB AS so
		ON mac.ORDER_NO = so.ORDER_NO AND mac.CONTRACT = so.CONTRACT
	WHERE
		so.CONTRACT = '2051'
		AND mac.ROWSTATE IN ('Released','Reserved')
	ORDER BY so.REVISED_START_DATE, mac.ORDER_NO
	"""
    return execute_query(query)

@timing_decorator("Load IPIS Data")
def load_ipis_data():
    """Load stock data from IPIS table"""
    query = """
    SELECT
		ipt.PART_NO,
		ipt.LOCATION_NO,
		ipt.LOT_BATCH_NO,
		ipt.QTY_ONHAND - ipt.QTY_RESERVED AS 'Available Qty'
	FROM IFS.INVENTORY_PART_IN_STOCK_TAB AS ipt
	WHERE 
		ipt.CONTRACT = '2051'
		AND ipt.WAREHOUSE <> 'QUALITY'
		AND ipt.AVAILABILITY_CONTROL_ID IS NULL
	UNION
	SELECT
		ipt.PART_NO,
		ipt.LOCATION_NO,
		ipt.LOT_BATCH_NO,
		ipt.QTY_ONHAND - ipt.QTY_RESERVED AS 'Available Qty'
	FROM IFS.INVENTORY_PART_IN_STOCK_TAB AS ipt
	WHERE 
		ipt.CONTRACT = '2051'
		AND ipt.WAREHOUSE <> 'QUALITY'
		AND ipt.AVAILABILITY_CONTROL_ID IN ('GOODS-INWARDS')
	"""
    return execute_query(query)

@timing_decorator("Load Hours Data")
def load_hours_data():
    """Load labor standards data from database"""
    query = """
    SELECT
		hrs.PART_NO,
		hrs.LABOR_CLASS_NO,
		CASE
			WHEN hrs.RUN_TIME_CODE IN (1,3) THEN hrs.LABOR_RUN_FACTOR
			WHEN hrs.RUN_TIME_CODE IN (2) THEN ISNULL(1/NULLIF(hrs.LABOR_RUN_FACTOR,0),0)
			ELSE 'Error' END AS 'Hours per Unit',
		'Kits' AS 'Area'
	FROM IFS.ROUTING_OPERATION_TAB AS hrs
	WHERE CONTRACT = '2051'
		AND PHASE_OUT_DATE IS NULL
		AND LABOR_CLASS_NO IN ('4936', '4948')
	UNION
	SELECT
		hrs.PART_NO,
		hrs.LABOR_CLASS_NO,
		CASE
			WHEN hrs.RUN_TIME_CODE IN (1,3) THEN hrs.LABOR_RUN_FACTOR
			WHEN hrs.RUN_TIME_CODE IN (2) THEN ISNULL(1/NULLIF(hrs.LABOR_RUN_FACTOR,0),0)
			ELSE 'Error' END AS 'Hours per Unit',
		CASE
			WHEN hrs.LABOR_CLASS_NO IN ('4931') THEN 'Manufacturing'
			WHEN hrs.LABOR_CLASS_NO IN ('4940') THEN 'Assembly'
			WHEN hrs.LABOR_CLASS_NO IN ('4941', '4947') THEN 'Packaging'
			WHEN hrs.LABOR_CLASS_NO IN ('4942') THEN 'Boxing'
			ELSE 'N/A' END AS 'Area'
	FROM IFS.ROUTING_OPERATION_TAB AS hrs
	WHERE CONTRACT = '2051'
	AND PHASE_OUT_DATE IS NULL
	AND LABOR_CLASS_NO IN ('4931', '4940', '4941', '4942', '4947');
    """
    return execute_query(query)

@timing_decorator("Load POs Data")
def load_pos_data():
    """Load purchase orders data from database"""
    query = """
    DECLARE @POWeeksOut INT;
	SET @POWeeksOut = 7;
	SELECT
		pol.ORDER_NO AS 'PO Number',
		pol.PART_NO AS 'Part Number',
		pol.BUY_QTY_DUE * pol.CONV_FACTOR AS 'Qty Due',
		pol.INVOICING_SUPPLIER AS 'Supplier',
		pol.LINE_NO AS 'PO Line',
		pol.PROMISED_DELIVERY_DATE AS 'Promised Due Date'
	FROM IFS.PURCHASE_ORDER_LINE_TAB AS pol
	INNER JOIN IFS.PURCHASE_ORDER_TAB AS po
		ON pol.ORDER_NO = po.ORDER_NO
	WHERE pol.PURCHASE_SITE = '2051'
		AND pol.ROWSTATE = 'Confirmed'
		AND pol.PROMISED_DELIVERY_DATE <= DATEADD(DAY, (@POWeeksOut+1)*7 - DATEPART(WEEKDAY, GETDATE()), GETDATE())
		AND pol.INVOICING_SUPPLIER NOT IN ('1060')
		ORDER BY pol.PROMISED_DELIVERY_DATE
    """
    return execute_query(query)

def test_database_connection():
    """Test the database connection and return status"""
    try:
        engine = get_database_connection()
        if engine:
            # Test the connection by executing a simple query
            with engine.connect() as connection:
                connection.execute(text("SELECT 1"))
            engine.dispose()
            return True, "Database connection successful"
    except Exception as e:
        return False, f"Database connection failed: {str(e)}"
    return False, "Database connection failed: Unknown error"

def get_table_info():
    """Get information about available tables in SQL Server"""
    try:
        engine = get_database_connection()
        with engine.connect() as connection:
            result = connection.execute(text("""
                SELECT table_name 
                FROM information_schema.tables 
                WHERE table_type = 'BASE TABLE'
            """))
            tables = [row[0] for row in result.fetchall()]
        engine.dispose()
        return tables
    except Exception as e:
        return []

def normalize_so_number(so_val):
    """Normalize SO numbers to handle Excel float formatting issues"""
    if pd.isna(so_val):
        return ""
    so_str = str(so_val).strip()
    # Remove trailing .0 if it's a whole number
    if so_str.endswith('.0') and so_str.replace('.0', '').isdigit():
        return so_str.replace('.0', '')
    return so_str

def safe_metric(metrics, key, default=0):
    """Safely get a metric value with a default if missing"""
    return metrics.get(key, default)

def format_metric(value, format_type='number'):
    """Format metric values consistently"""
    try:
        if format_type == 'number':
            return f"{value:,}"
        elif format_type == 'hours':
            return f"{value:,.1f}"
        elif format_type == 'percentage':
            return f"{value:.1f}%"
        return str(value)
    except (TypeError, ValueError):
        return "0"  # Safe default for invalid values

def get_sorting_strategies():
    """Define all sorting strategies for min/max optimization"""
    return [
        {"name": "Start Date (Early First)", "columns": ["Start Date", "SO Number"], "ascending": [True, True]},
        {"name": "Start Date (Late First)", "columns": ["Start Date", "SO Number"], "ascending": [False, True]},
        {"name": "Demand (Small First)", "columns": ["Demand", "Start Date"], "ascending": [True, True]},
        {"name": "Demand (Large First)", "columns": ["Demand", "Start Date"], "ascending": [False, True]},
        {"name": "Hours (Quick First)", "columns": ["Hours_Calc", "Start Date"], "ascending": [True, True]},
        {"name": "Hours (Long First)", "columns": ["Hours_Calc", "Start Date"], "ascending": [False, True]},
        {"name": "Part Number (A-Z)", "columns": ["Part", "Start Date"], "ascending": [True, True]},
        {"name": "Part Number (Z-A)", "columns": ["Part", "Start Date"], "ascending": [False, True]},
        {"name": "Planner (A-Z)", "columns": ["Planner", "Start Date"], "ascending": [True, True]},
        {"name": "Planner (Z-A)", "columns": ["Planner", "Start Date"], "ascending": [False, True]}
    ]

@timing_decorator("Build Stock Dictionary")
def build_stock_dictionary(df_ipis):
    """Build stock dict using IPIS as primary source"""
    stock = {}
    
    # Use IPIS as the authoritative source
    if not df_ipis.empty:
        df_ipis["PART_NO"] = df_ipis["PART_NO"].astype(str)
        ipis_stock = df_ipis.groupby("PART_NO")["Available Qty"].sum().to_dict()
        stock.update(ipis_stock)
    else:
        print("WARNING: IPIS sheet is empty - no stock data available!")
    
    return stock

def process_single_scenario(scenario_name, status_callback=None, scenario_num=1, total_scenarios=1, sorting_strategy=None, include_kits=True, include_instruments=True, include_virtuoso=True, include_kit_samples=True):
    """Process a single scenario from database and return results with live progress updates"""
    
    # Start overall processing timing
    performance_tracker.start_phase("Total Processing")
    
    if status_callback:
        strategy_name = sorting_strategy["name"] if sorting_strategy else "Default"
        status_callback(f"📂 [Scenario {scenario_num}/{total_scenarios}] Loading data from database ({strategy_name})...")
    
    # Load the data from database with proper timing separation
    try:
        # Test database connection first (this will be included in the data loading timing)
        performance_tracker.start_phase("Database Connection")
        test_connection = get_database_connection()
        performance_tracker.end_phase()
        
        # Load all data tables
        df_main = load_demand_data()
        df_struct = load_planned_demand_data()
        df_component_demand = load_component_demand_data()
        df_ipis = load_ipis_data()
        df_hours = load_hours_data()
        df_pos = load_pos_data()
    except Exception as e:
        performance_tracker.end_phase()  # End total processing
        raise Exception(f"Failed to load data from database: {str(e)}")
    
    # DEBUG: Show data loading results
    if DEBUG_MODE:
        print(f"\n🔍 DEBUG: Data loaded from database")
        print(f"   Demand table: {len(df_main)} rows")
        print(f"   Planned Demand table: {len(df_struct)} rows")
        print(f"   Component Demand table: {len(df_component_demand)} rows")
        print(f"   IPIS table: {len(df_ipis)} rows")
        print(f"   Hours table: {len(df_hours)} rows")
        print(f"   POs table: {len(df_pos)} rows")
        
        # Show sample data from each table to verify data quality
        if DEBUG_COMPONENT_PART is not None:
            debug_part = str(DEBUG_COMPONENT_PART)
            print(f"\n🔍 DEBUG: Data quality check for component {debug_part}")
            
            # Check IPIS (stock) data
            print(f"\n  📦 IPIS (Stock) Data for {debug_part}:")
            ipis_matches = df_ipis[df_ipis["PART_NO"].astype(str) == debug_part]
            if not ipis_matches.empty:
                for _, row in ipis_matches.iterrows():
                    print(f"    PART_NO: {row['PART_NO']}, Available Qty: {row['Available Qty']}")
            else:
                print(f"    ❌ No stock data found for {debug_part}")
            
            # Check Component Demand (committed) data
            print(f"\n  🔒 Component Demand (Committed) Data for {debug_part}:")
            committed_matches = df_component_demand[df_component_demand["Component Part Number"].astype(str) == debug_part]
            if not committed_matches.empty:
                for _, row in committed_matches.iterrows():
                    print(f"    Component: {row['Component Part Number']}, Qty Required: {row['Component Qty Required']}")
            else:
                print(f"    ❌ No committed demand found for {debug_part}")
            
            # Check Planned Demand (BOM) data
            print(f"\n  📋 Planned Demand (BOM) Data for {debug_part}:")
            planned_matches = df_struct[df_struct["Component Part Number"].astype(str) == debug_part]
            if not planned_matches.empty:
                for _, row in planned_matches.iterrows():
                    print(f"    SO: {row['SO Number']}, Component: {row['Component Part Number']}, Qty: {row['Component Qty Required']}")
            else:
                print(f"    ❌ No BOM data found for {debug_part}")
            
            # Check POs data
            print(f"\n  📄 POs Data for {debug_part}:")
            po_matches = df_pos[df_pos["Part Number"].astype(str) == debug_part]
            if not po_matches.empty:
                for _, row in po_matches.iterrows():
                    print(f"    PO: {row['PO Number']}, Part: {row['Part Number']}, Qty: {row['Qty Due']}, Due: {row['Promised Due Date']}")
            else:
                print(f"    ❌ No PO data found for {debug_part}")
            
            # Check Hours data
            print(f"\n  ⏱️ Hours Data for {debug_part}:")
            hours_matches = df_hours[df_hours["PART_NO"].astype(str) == debug_part]
            if not hours_matches.empty:
                for _, row in hours_matches.iterrows():
                    print(f"    Part: {row['PART_NO']}, Hours per Unit: {row['Hours per Unit']}")
            else:
                print(f"    ❌ No hours data found for {debug_part}")
            
            print("-" * 80)
    
    # Validate required columns in Demand table
    required_columns = ['SO No', 'Part No', 'Planner', 'Start Date', 'Rev Qty Due']
    missing_columns = [col for col in required_columns if col not in df_main.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns in Demand table: {', '.join(missing_columns)}")
    
    # Rename columns to match the rest of the code's expectations
    df_main = df_main.rename(columns={
        'SO No': 'SO Number',
        'Part No': 'Part',
        'Rev Qty Due': 'Demand'
    })
    
    if status_callback:
        strategy_name = sorting_strategy["name"] if sorting_strategy else "Default"
        status_callback(f"🔁 [Scenario {scenario_num}/{total_scenarios}] Processing commitments ({strategy_name})...")
    
    # Build stock dictionary
    stock = build_stock_dictionary(df_ipis)
    
    # DEBUG: Show stock information
    if DEBUG_MODE:
        print(f"\n🔍 DEBUG: Stock information")
        print(f"   Total stock entries: {len(stock)}")
        if DEBUG_COMPONENT_PART and str(DEBUG_COMPONENT_PART) in stock:
            print(f"   🎯 DEBUG PART {DEBUG_COMPONENT_PART}: {stock[str(DEBUG_COMPONENT_PART)]} available")
        
        # Show detailed stock building process for debug component
        if DEBUG_COMPONENT_PART is not None:
            debug_part = str(DEBUG_COMPONENT_PART)
            print(f"\n🔍 DEBUG: Stock building process for {debug_part}")
            
            # Show raw IPIS data
            print(f"  📦 Raw IPIS data for {debug_part}:")
            ipis_raw = df_ipis[df_ipis["PART_NO"].astype(str) == debug_part]
            if not ipis_raw.empty:
                total_stock = 0
                for _, row in ipis_raw.iterrows():
                    qty = row['Available Qty']
                    total_stock += qty
                    print(f"    Row: PART_NO='{row['PART_NO']}', Available Qty={qty}")
                print(f"    Total stock from IPIS: {total_stock}")
                print(f"    Final stock in dictionary: {stock.get(debug_part, 0)}")
            else:
                print(f"    ❌ No IPIS rows found for {debug_part}")
                print(f"    Final stock in dictionary: {stock.get(debug_part, 0)}")
            
            print("-" * 80)

    # Build committed_components
    committed_components = {}
    committed_parts_count = 0
    total_committed_qty = 0

    if not df_component_demand.empty:
        df_component_demand["Component Part Number"] = df_component_demand["Component Part Number"].astype(str)
        committed_summary = df_component_demand.groupby("Component Part Number")["Component Qty Required"].sum()
        committed_components = committed_summary.to_dict()
        committed_parts_count = len(committed_components)
        total_committed_qty = sum(committed_components.values())
        
        # DEBUG: Show committed components
        if DEBUG_MODE:
            print(f"\n🔍 DEBUG: Committed components")
            print(f"   Total committed parts: {committed_parts_count}")
            print(f"   Total committed quantity: {total_committed_qty}")
            if DEBUG_COMPONENT_PART and str(DEBUG_COMPONENT_PART) in committed_components:
                print(f"   🎯 DEBUG PART {DEBUG_COMPONENT_PART}: {committed_components[str(DEBUG_COMPONENT_PART)]} committed")
        
        # Show detailed committed components building process for debug component
        if DEBUG_COMPONENT_PART is not None:
            debug_part = str(DEBUG_COMPONENT_PART)
            print(f"\n🔍 DEBUG: Committed components building process for {debug_part}")
            
            # Show raw Component Demand data
            print(f"  🔒 Raw Component Demand data for {debug_part}:")
            committed_raw = df_component_demand[df_component_demand["Component Part Number"].astype(str) == debug_part]
            if not committed_raw.empty:
                total_committed = 0
                for _, row in committed_raw.iterrows():
                    qty = row['Component Qty Required']
                    total_committed += qty
                    print(f"    Row: Component='{row['Component Part Number']}', Qty Required={qty}")
                print(f"    Total committed from Component Demand: {total_committed}")
                print(f"    Final committed in dictionary: {committed_components.get(debug_part, 0)}")
            else:
                print(f"    ❌ No Component Demand rows found for {debug_part}")
                print(f"    Final committed in dictionary: {committed_components.get(debug_part, 0)}")
            
            print("-" * 80)

    # Initialize used_components with the committed quantities
    used_components = committed_components.copy()

    # Build labor standards dictionary (unchanged)
    df_hours["PART_NO"] = df_hours["PART_NO"].astype(str)
    labor_standards = df_hours.groupby("PART_NO")["Hours per Unit"].sum().to_dict()
    
    if status_callback:
        strategy_name = sorting_strategy["name"] if sorting_strategy else "Default"
        status_callback(f"🔁 [Scenario {scenario_num}/{total_scenarios}] Building planned demand structures ({strategy_name})...")
    
    # Build planned demand structure
    planned_demand = df_struct[df_struct["Component Part Number"].notna()].copy()
    planned_demand["SO Number"] = planned_demand["SO Number"].astype(str).str.strip()
    planned_demand["Component Part Number"] = planned_demand["Component Part Number"].astype(str)
    
    # Apply normalization to both planned demand and main data
    planned_demand["SO Number"] = planned_demand["SO Number"].apply(normalize_so_number)
    
    # Pre-process main data (unchanged)
    df_main['Start Date'] = pd.to_datetime(df_main['Start Date'], errors='coerce')
    df_main["Part"] = df_main["Part"].astype(str)
    df_main["Planner"] = df_main["Planner"].fillna("UNKNOWN").astype(str)
    df_main["Demand"] = pd.to_numeric(df_main["Demand"], errors='coerce').fillna(0)
    
    # Filter data based on selected categories
    filtered_df_main = df_main.copy()
    
    # Define planner codes for each category
    kits_planners = ['3001', '3801', '5001']  # BVI Kits (3001, 3801) + Malosa Kits (5001)
    instruments_planners = ['3802', '3803', '3804', '3805']  # Manufacturing, Assembly, Packaging, Malosa Instruments
    virtuoso_planners = ['3806']  # Virtuoso
    kit_samples_planners = ['KIT SAMPLES']
    
    # Build filter mask based on selected categories
    filter_mask = pd.Series([False] * len(df_main), index=df_main.index)
    
    if include_kits:
        filter_mask |= df_main['Planner'].isin(kits_planners)
    
    if include_instruments:
        filter_mask |= df_main['Planner'].isin(instruments_planners)
    
    if include_virtuoso:
        filter_mask |= df_main['Planner'].isin(virtuoso_planners)
    
    if include_kit_samples:
        filter_mask |= df_main['Planner'].isin(kit_samples_planners)
    
    # Apply filter
    filtered_df_main = df_main[filter_mask].copy()
    
    if status_callback:
        total_original = len(df_main)
        total_filtered = len(filtered_df_main)
        excluded = total_original - total_filtered
        status_callback(f"🔁 [Scenario {scenario_num}/{total_scenarios}] Filtered data: {total_filtered:,}/{total_original:,} orders selected ({excluded:,} excluded) ({strategy_name})...")
    
    # Calculate hours for sorting (unchanged)
    filtered_df_main["Hours_Calc"] = filtered_df_main.apply(lambda row: 
        labor_standards.get(str(row["Part"]), 0) * row["Demand"], axis=1)
    
    # Apply sorting strategy (unchanged)
    if sorting_strategy:
        # Handle missing values appropriately for each column type
        for col in sorting_strategy["columns"]:
            if col == "Start Date":
                # Put NaT (missing dates) at the end
                filtered_df_main = filtered_df_main.sort_values(sorting_strategy["columns"], 
                                            ascending=sorting_strategy["ascending"], 
                                            na_position='last')
            else:
                filtered_df_main = filtered_df_main.sort_values(sorting_strategy["columns"], 
                                            ascending=sorting_strategy["ascending"])
    else:
        # Default sorting (original behavior)
        filtered_df_main = filtered_df_main.sort_values(['Start Date', 'SO Number'], na_position='last')
    
    filtered_df_main = filtered_df_main.reset_index(drop=True)
    
    results = []
    # Start order processing timing
    performance_tracker.start_phase("Order Processing")
    
    processed = 0
    total = len(filtered_df_main)
    
    # Baseline estimate: ~0.15 seconds per order (conservative estimate)
    baseline_time_per_order = 0.15
    processing_start_time = time.time()

    # Process each order sequentially with FREQUENT UI updates + TIME ESTIMATES
    
    # DEBUG: Show header when starting to process orders
    debug_component_orders_found = 0
    if DEBUG_MODE and DEBUG_COMPONENT_PART is not None:
        print(f"\n🎯 DEBUG: Starting to process orders that use component {DEBUG_COMPONENT_PART}")
        print("="*80)
    
    for _, row in filtered_df_main.iterrows():
        processed += 1
        
        # UPDATE UI EVERY 100 ORDERS
        if processed % 100 == 0 or processed == total or processed == 1:
            progress_pct = processed / total * 100
            
            # Calculate dynamic time estimates
            if processed >= 10:  # After 10 orders, use actual performance
                elapsed = time.time() - processing_start_time
                actual_time_per_order = elapsed / processed
                remaining_orders = total - processed
                est_remaining = remaining_orders * actual_time_per_order
            else:  # For first few orders, use baseline estimate
                elapsed = time.time() - processing_start_time
                remaining_orders = total - processed
                est_remaining = remaining_orders * baseline_time_per_order
            

            
            if status_callback:
                strategy_name = sorting_strategy["name"] if sorting_strategy else "Default"
                if processed == total:
                    status_callback(f"✅ [Scenario {scenario_num}/{total_scenarios}] Database ({strategy_name}) - Completed {total:,} orders in {elapsed:.1f}s")
                else:
                    # Show current scenario progress + context about remaining scenarios
                    remaining_scenarios = total_scenarios - scenario_num
                    if remaining_scenarios > 0:
                        status_callback(f"🔁 [Scenario {scenario_num}/{total_scenarios}] {strategy_name} - {processed:,}/{total:,} ({progress_pct:.1f}%) | {est_remaining:.0f}s + {remaining_scenarios} more")
                    else:
                        status_callback(f"🔁 [Scenario {scenario_num}/{total_scenarios}] {strategy_name} - {processed:,}/{total:,} ({progress_pct:.1f}%) | {est_remaining:.0f}s remaining")
        
        so = str(row["SO Number"]).strip() if pd.notna(row["SO Number"]) else f"ORDER_{processed}"
        part = str(row["Part"]) if pd.notna(row["Part"]) else None
        demand_qty = row["Demand"] if pd.notna(row["Demand"]) and row["Demand"] > 0 else 0
        planner = str(row["Planner"]) if pd.notna(row["Planner"]) else "UNKNOWN"
        start_date = row["Start Date"]
        
        # NORMALIZE SO NUMBER for consistent matching
        so = normalize_so_number(so)
        
        # DEBUG: Show when processing specific SO or part
        debug_so_match = DEBUG_SO_NUMBER is not None and str(so) == str(DEBUG_SO_NUMBER)
        debug_part_match = DEBUG_COMPONENT_PART is not None and str(part) == str(DEBUG_COMPONENT_PART)
        
        if DEBUG_MODE and (debug_so_match or debug_part_match):
            debug_component_orders_found += 1
            print(f"\n🎯 DEBUG: Processing order {processed}/{total} (Order #{debug_component_orders_found} using {DEBUG_COMPONENT_PART})")
            print(f"   SO: {so}")
            print(f"   Part: {part}")
            print(f"   Demand: {demand_qty}")
            print(f"   Planner: {planner}")
            print(f"   Start Date: {start_date}")
        
        # ENHANCED DEBUG: Track every Shop Order that tries to allocate to the debug component
        debug_component_allocation_found = False
        if DEBUG_MODE and DEBUG_COMPONENT_PART is not None:
            # Check if this SO will try to allocate the debug component (either as parent part or as component)
            debug_component = str(DEBUG_COMPONENT_PART)
            if str(part) == debug_component:
                debug_component_allocation_found = True
                print(f"\n🔍 DEBUG COMPONENT ALLOCATION: SO {so} directly uses {debug_component} as parent part")
            else:
                # Check if this SO's BOM contains the debug component
                try:
                    bom_check = planned_demand[planned_demand["SO Number"] == so]
                    if not bom_check.empty:
                        bom_components = bom_check["Component Part Number"].astype(str).tolist()
                        if debug_component in bom_components:
                            debug_component_allocation_found = True
                            comp_row = bom_check[bom_check["Component Part Number"].astype(str) == debug_component].iloc[0]
                            comp_qty = comp_row["Component Qty Required"]
                            print(f"\n🔍 DEBUG COMPONENT ALLOCATION: SO {so} (Parent: {part}) requires {comp_qty} units of {debug_component}")
                except:
                    pass
        
        # Skip orders with missing critical data
        if part is None or part == "nan" or demand_qty <= 0:
            # DEBUG: Show skipped orders
            if DEBUG_MODE and (debug_so_match or debug_part_match):
                print(f"   ⚠️  Order skipped: part={part}, demand={demand_qty}")
            
            results.append({
                "SO Number": so,
                "Part": part or "MISSING",
                "Planner": planner,
                "Start Date": start_date.strftime('%Y-%m-%d') if pd.notna(start_date) else "No Date",
                "PB": "-",
                "Demand": demand_qty,
                "Hours": 0,
                "Status": "⚠️ Skipped",
                "Shortages": "-",
                "Components": "Missing part number or zero demand"
            })
            continue
        
        # Check if this is a piggyback order
        try:
            pb_check = f"NS{part}99"
            is_pb = "PB" if pb_check in planned_demand["Component Part Number"].values else "-"
        except:
            is_pb = "-"
        
        # Get planned demand for this SO
        try:
            bom = planned_demand[planned_demand["SO Number"] == so]
        except:
            bom = pd.DataFrame()
        
        # DEBUG: Show BOM lookup results for specific SO
        debug_so_match = DEBUG_SO_NUMBER is not None and str(so) == str(DEBUG_SO_NUMBER)
        if DEBUG_MODE and debug_so_match:
            print(f"\n=== DEBUG: BOM Lookup for SO {so} ===")
            print(f"  Total planned demand records: {len(planned_demand)}")
            print(f"  Looking for SO Number: '{so}' (type: {type(so)}, repr: {repr(so)})")
            print(f"  Note: SO Number has been normalized to handle Excel formatting issues")
            
            # Show unique SO numbers in planned demand to see the format
            unique_so_numbers = planned_demand["SO Number"].unique()
            print(f"  Sample SO numbers in planned demand (normalized): {unique_so_numbers[:10].tolist()}")
            
            # Check if our SO exists in any form
            so_found = planned_demand["SO Number"].astype(str).str.contains(str(so), na=False)
            print(f"  Records containing '{so}' (anywhere): {so_found.sum()}")
            
            # ENHANCED DEBUG: Show the actual records that contain the SO number
            if so_found.sum() > 0:
                print(f"\n  🔍 RECORDS CONTAINING '{so}' (showing raw data):")
                matching_records = planned_demand[so_found]
                for idx, row in matching_records.iterrows():
                    raw_so = row["SO Number"]
                    print(f"    Row {idx}: SO='{raw_so}' (type: {type(raw_so)}, repr: {repr(raw_so)})")
                    print(f"           Component='{row['Component Part Number']}', Qty={row['Component Qty Required']}")
            
            print(f"  Records matching SO {so} exactly: {len(bom)}")
            if len(bom) > 0:
                print(f"  Components found:")
                for _, comp in bom.iterrows():
                    comp_part = str(comp["Component Part Number"])
                    comp_qty = comp["Component Qty Required"]
                    print(f"    {comp_part}: {comp_qty}")
            else:
                print(f"  No components found - treating as raw material")
                # Show a few sample records from planned demand
                print(f"  Sample planned demand records:")
                for i, (_, row) in enumerate(planned_demand.head(5).iterrows()):
                    raw_so = row["SO Number"]
                    print(f"    Row {i}: SO='{raw_so}' (type: {type(raw_so)}, repr: {repr(raw_so)}), Component='{row['Component Part Number']}', Qty={row['Component Qty Required']}")
            print("-" * 80)
        
        # ENHANCED DEBUG: Show BOM lookup for any SO that uses the debug component
        if DEBUG_MODE and DEBUG_COMPONENT_PART is not None:
            debug_part = str(DEBUG_COMPONENT_PART)
            debug_so_match = DEBUG_SO_NUMBER is not None and str(so) == str(DEBUG_SO_NUMBER)
            debug_part_match = str(part) == debug_part
            
            if debug_so_match or debug_part_match:
                print(f"\n=== DEBUG: BOM Lookup for SO {so} (Parent: {part}) ===")
                print(f"  Looking for SO Number: '{so}' in planned demand")
                print(f"  Records matching SO {so} exactly: {len(bom)}")
                
                if len(bom) > 0:
                    print(f"  Components found in BOM:")
                    for _, comp in bom.iterrows():
                        comp_part = str(comp["Component Part Number"])
                        comp_qty = comp["Component Qty Required"]
                        is_debug_component = comp_part == debug_part
                        debug_marker = " 🎯" if is_debug_component else ""
                        print(f"    {comp_part}: {comp_qty}{debug_marker}")
                else:
                    print(f"  No BOM found - treating as raw material")
                    if debug_part_match:
                        print(f"  🎯 This SO directly uses {debug_part} as parent part")
                
                print("-" * 80)
        
        # Check material availability
        releasable = True
        shortage_details = []
        components_needed = {}
        
        # Calculate labor hours for this order
        base_hours = labor_standards.get(part, 0)
        labor_hours = base_hours * demand_qty
        
        if len(bom) > 0:
            # This SO has planned component demand - use ALL-OR-NOTHING allocation
            all_components_available = True
            component_requirements = []
            debug_so_match = DEBUG_SO_NUMBER is not None and str(so) == str(DEBUG_SO_NUMBER)
            if DEBUG_MODE and debug_so_match:
                print(f"\n=== DEBUG: Processing SO {so} ===")
                print(f"Found {len(bom)} components in planned demand")
            for _, comp in bom.iterrows():
                try:
                    comp_part = str(comp["Component Part Number"])
                    required_qty = int(comp["Component Qty Required"]) if pd.notna(comp["Component Qty Required"]) else 0
                    total_used = used_components.get(comp_part, 0)
                    true_available = stock.get(comp_part, 0) - total_used
                    available = max(0, true_available)

                    # Always use the 'what if' calculation for both debug and output
                    available_after_usage = true_available - required_qty
                    will_be_sufficient = true_available >= required_qty

                    # Debug output for selected component part and/or SO number
                    debug_part_match = DEBUG_COMPONENT_PART is not None and str(comp_part) == str(DEBUG_COMPONENT_PART)
                    # Note: debug_so_match is already set above for BOM processing
                    
                    # FIX: Always print debug for every component in BOM if SO matches
                    if DEBUG_MODE and (debug_part_match or debug_so_match):
                        print(f"\n=== DEBUG: SO {so} (Parent: {part}) requires component {comp_part} ===")
                        print(f"  📊 STOCK ALLOCATION CALCULATION for component {comp_part}:")
                        print(f"    Initial stock:           {stock.get(comp_part, 0):>8}")
                        print(f"    Committed qty:           {committed_components.get(comp_part, 0) if 'committed_components' in locals() else 0:>8}")
                        print(f"    Already allocated:       {total_used:>8}")
                        print(f"    Available for this SO:   {true_available:>8}")
                        print(f"")
                        print(f"    Required for SO {so}:    {required_qty:>8}")
                        print(f"    Would remain after:      {available_after_usage:>8}")
                        print(f"")
                        print(f"    ✅ CAN FULFILL ORDER:    {will_be_sufficient}")
                        if will_be_sufficient:
                            print(f"    📦 ALLOCATION: {required_qty} units allocated to SO {so}")
                            print(f"    📦 REMAINING: {available_after_usage} units left in stock")
                        else:
                            print(f"    ❌ SHORTAGE: Need {required_qty}, have {true_available}, short {abs(available_after_usage)}")
                        
                        # Show detailed calculation breakdown
                        print(f"\n  🔍 DETAILED CALCULATION BREAKDOWN:")
                        print(f"    Stock lookup: stock.get('{comp_part}', 0) = {stock.get(comp_part, 0)}")
                        print(f"    Used lookup: used_components.get('{comp_part}', 0) = {total_used}")
                        print(f"    Calculation: {stock.get(comp_part, 0)} - {total_used} = {true_available}")
                        print(f"    Required: {required_qty}")
                        print(f"    Sufficient: {true_available} >= {required_qty} = {will_be_sufficient}")
                        
                        # Show all POs for this part
                        future_pos = df_pos[
                            (df_pos['Part Number'].astype(str) == str(comp_part)) &
                            (pd.to_datetime(df_pos['Promised Due Date'], errors='coerce') >= datetime.now())
                        ]
                        if not future_pos.empty:
                            print(f"\n  📋 Future POs for {comp_part}:")
                            for _, po_row in future_pos.iterrows():
                                po_id = po_row['PO Number']
                                po_qty = po_row['Qty Due']
                                po_date = pd.to_datetime(po_row['Promised Due Date']).strftime('%Y-%m-%d')
                                print(f"    PO {po_id}: {po_qty} due {po_date}")
                        else:
                            print(f"\n  📋 No future POs found for {comp_part}")
                        print("-" * 80)

                    component_requirements.append({
                        'part': comp_part,
                        'required': required_qty,
                        'available': available
                    })
                    components_needed[comp_part] = required_qty

                    # Check availability but DON'T allocate yet
                    if not will_be_sufficient:
                        all_components_available = False
                        shortage = abs(available_after_usage)  # Changed to use available_after_usage directly
                        # Search POs for potential resolution
                        future_pos = df_pos[
                            (df_pos["Part Number"].astype(str) == comp_part) &
                            (pd.to_datetime(df_pos["Promised Due Date"], errors="coerce") >= datetime.now())
                        ]
                        po_match = None
                        for _, po_row in future_pos.iterrows():
                            po_qty = po_row["Qty Due"]
                            if po_qty >= shortage:
                                po_match = po_row
                                break
                        if po_match is not None:
                            po_id = po_match["PO Number"]
                            po_date = pd.to_datetime(po_match["Promised Due Date"]).strftime('%Y-%m-%d')
                            shortage_details.append(f"{comp_part} short {shortage} – PO {po_id} due {po_date}")
                        else:
                            shortage_details.append(f"{comp_part} (need {required_qty}, have {true_available}, short {shortage})")
                except Exception as e:
                    all_components_available = False
                    shortage_details.append(f"Component processing error: {str(e)}")
                    continue

            # Only after all checks, allocate components if all are available
            if all_components_available:
                releasable = True
                for req in component_requirements:
                    comp_part = req['part']
                    required_qty = req['required']
                    used_components[comp_part] = used_components.get(comp_part, 0) + required_qty
            else:
                releasable = False

        elif len(bom) == 0:
            # This SO has no planned component demand - treat as raw material/purchased part
            try:
                total_used = used_components.get(part, 0)
                true_available = stock.get(part, 0) - total_used  # Changed to use true_available
                available_after_usage = true_available - demand_qty  # Added to match debug logic
                
                # DEBUG: Show allocation calculation for raw material parts (only for raw material orders)
                debug_part_match = DEBUG_COMPONENT_PART is not None and str(part) == str(DEBUG_COMPONENT_PART)
                debug_so_match = DEBUG_SO_NUMBER is not None and str(so) == str(DEBUG_SO_NUMBER)
                
                if DEBUG_MODE and (debug_part_match or debug_so_match):
                    print(f"\n=== DEBUG: Raw Material {part} for SO {so} ===")
                    print(f"  📊 STOCK ALLOCATION CALCULATION:")
                    print(f"    Initial stock:           {stock.get(part, 0):>8}")
                    print(f"    Committed qty:           {committed_components.get(part, 0) if 'committed_components' in locals() else 0:>8}")
                    print(f"    Already allocated:       {total_used:>8}")
                    print(f"    Available for this SO:   {true_available:>8}")
                    print(f"")
                    print(f"    Required for SO {so}:    {demand_qty:>8}")
                    print(f"    Would remain after:      {available_after_usage:>8}")
                    print(f"")
                    print(f"    ✅ CAN FULFILL ORDER:    {true_available >= demand_qty}")
                    
                    if true_available >= demand_qty:
                        print(f"    📦 ALLOCATION: {demand_qty} units allocated to SO {so}")
                        print(f"    📦 REMAINING: {available_after_usage} units left in stock")
                    else:
                        print(f"    ❌ SHORTAGE: Need {demand_qty}, have {true_available}, short {abs(available_after_usage)}")
                    
                    # Show detailed calculation breakdown for raw material
                    print(f"\n  🔍 DETAILED CALCULATION BREAKDOWN:")
                    print(f"    Stock lookup: stock.get('{part}', 0) = {stock.get(part, 0)}")
                    print(f"    Used lookup: used_components.get('{part}', 0) = {total_used}")
                    print(f"    Calculation: {stock.get(part, 0)} - {total_used} = {true_available}")
                    print(f"    Required: {demand_qty}")
                    print(f"    Sufficient: {true_available} >= {demand_qty} = {true_available >= demand_qty}")
                    
                    # Show all POs for this part
                    future_pos = df_pos[
                        (df_pos['Part Number'].astype(str) == str(part)) &
                        (pd.to_datetime(df_pos['Promised Due Date'], errors='coerce') >= datetime.now())
                    ]
                    if not future_pos.empty:
                        print(f"\n  📋 Future POs for {part}:")
                        for _, po_row in future_pos.iterrows():
                            po_id = po_row['PO Number']
                            po_qty = po_row['Qty Due']
                            po_date = pd.to_datetime(po_row['Promised Due Date']).strftime('%Y-%m-%d')
                            print(f"    PO {po_id}: {po_qty} due {po_date}")
                    else:
                        print(f"\n  📋 No future POs found for {part}")
                    
                    print("-" * 80)
                
                if true_available >= demand_qty:  # Changed to use true_available
                    used_components[part] = used_components.get(part, 0) + demand_qty
                    releasable = True
                else:
                    releasable = False
                    shortage = abs(available_after_usage)  # Changed to use available_after_usage
                    shortage_details.append(f"{part} (need {demand_qty}, have {true_available}, short {shortage})")
            except:
                releasable = False
                shortage_details.append(f"{part} (stock lookup failed)")

        # Build result record
        shortage_parts_only = []
        components_info = "; ".join(shortage_details) if shortage_details else str(components_needed) if components_needed else "-"

        # Extract just the part numbers from shortage details
        for detail in shortage_details:
            if " short " in detail and "–" in detail:
                part_short = detail.split(" short ")[0].strip()
                shortage_parts_only.append(part_short)
            elif "(" in detail and " (need " in detail:
                part_short = detail.split(" (need ")[0].strip()
                shortage_parts_only.append(part_short)
            elif "(" in detail:
                part_short = detail.split("(")[0].strip()
                shortage_parts_only.append(part_short)
            else:
                part_short = detail.split()[0] if detail.split() else detail
                shortage_parts_only.append(part_short)

        clean_shortages = "; ".join(shortage_parts_only) if shortage_parts_only else "-"

        # DEBUG: Show final order decision
        if DEBUG_MODE and (debug_so_match or debug_part_match):
            print(f"\n  🎯 FINAL ORDER DECISION:")
            print(f"    SO {so} ({part}): {'✅ RELEASABLE' if releasable else '❌ ON HOLD'}")
            if not releasable and shortage_details:
                print(f"    Reason: {shortage_details[0] if shortage_details else 'Unknown'}")
            print(f"    Components needed: {components_needed if components_needed else 'None (raw material)'}")
            print("=" * 80)

        results.append({
            "SO Number": so,
            "Part": part,
            "Planner": planner,
            "Start Date": start_date.strftime('%Y-%m-%d') if pd.notna(start_date) else "No Date",
            "PB": is_pb,
            "Demand": demand_qty,
            "Hours": round(labor_hours, 4),
            "Status": "✅ Release" if releasable else "❌ Hold",
            "Shortages": clean_shortages,
            "Components": components_info
        })

    # Calculate summary metrics
    df_results = pd.DataFrame(results)
    total_orders = len(df_results)
    releasable_count = len(df_results[df_results['Status'] == '✅ Release'])
    held_count = total_orders - releasable_count
    pb_count = len(df_results[df_results['PB'] == 'PB'])
    skipped_count = len(df_results[df_results['Status'] == '⚠️ Skipped'])
    
    total_hours = df_results['Hours'].sum()
    releasable_hours = df_results[df_results['Status'] == '✅ Release']['Hours'].sum()
    held_hours = df_results[df_results['Status'] == '❌ Hold']['Hours'].sum()
    
    # Calculate quantity metrics
    total_qty = df_results['Demand'].sum()
    releasable_qty = df_results[df_results['Status'] == '✅ Release']['Demand'].sum()
    held_qty = df_results[df_results['Status'] == '❌ Hold']['Demand'].sum()
    
    # Calculate Kit and Instrument metrics with subcategories
    releasable_results = df_results[df_results['Status'] == '✅ Release']
    held_results = df_results[df_results['Status'] == '❌ Hold']
    
    # BVI Kits (Planner codes 3001, 3801)
    bvi_kit_planners = ['3001', '3801']
    total_bvi_kits = df_results[df_results['Planner'].isin(bvi_kit_planners)]
    releasable_bvi_kits = releasable_results[releasable_results['Planner'].isin(bvi_kit_planners)]
    total_bvi_kits_count = len(total_bvi_kits)
    total_bvi_kits_hours = total_bvi_kits['Hours'].sum()
    total_bvi_kits_qty = total_bvi_kits['Demand'].sum()
    releasable_bvi_kits_count = len(releasable_bvi_kits)
    releasable_bvi_kits_hours = releasable_bvi_kits['Hours'].sum()
    releasable_bvi_kits_qty = releasable_bvi_kits['Demand'].sum()
    
    # Malosa Kits (Planner code 5001)
    malosa_kit_planners = ['5001']
    total_malosa_kits = df_results[df_results['Planner'].isin(malosa_kit_planners)]
    releasable_malosa_kits = releasable_results[releasable_results['Planner'].isin(malosa_kit_planners)]
    total_malosa_kits_count = len(total_malosa_kits)
    total_malosa_kits_hours = total_malosa_kits['Hours'].sum()
    total_malosa_kits_qty = total_malosa_kits['Demand'].sum()
    releasable_malosa_kits_count = len(releasable_malosa_kits)
    releasable_malosa_kits_hours = releasable_malosa_kits['Hours'].sum()
    releasable_malosa_kits_qty = releasable_malosa_kits['Demand'].sum()
    
    # Total Kits
    total_kits_count = total_bvi_kits_count + total_malosa_kits_count
    total_kits_hours = total_bvi_kits_hours + total_malosa_kits_hours
    total_kits_qty = total_bvi_kits_qty + total_malosa_kits_qty
    releasable_kits_count = releasable_bvi_kits_count + releasable_malosa_kits_count
    releasable_kits_hours = releasable_bvi_kits_hours + releasable_malosa_kits_hours
    releasable_kits_qty = releasable_bvi_kits_qty + releasable_malosa_kits_qty
    
    # Manufacturing (Planner code 3802)
    manufacturing_planners = ['3802']
    total_manufacturing = df_results[df_results['Planner'].isin(manufacturing_planners)]
    releasable_manufacturing = releasable_results[releasable_results['Planner'].isin(manufacturing_planners)]
    total_manufacturing_count = len(total_manufacturing)
    total_manufacturing_hours = total_manufacturing['Hours'].sum()
    total_manufacturing_qty = total_manufacturing['Demand'].sum()
    releasable_manufacturing_count = len(releasable_manufacturing)
    releasable_manufacturing_hours = releasable_manufacturing['Hours'].sum()
    releasable_manufacturing_qty = releasable_manufacturing['Demand'].sum()
    
    # Assembly (Planner code 3803)
    assembly_planners = ['3803']
    total_assembly = df_results[df_results['Planner'].isin(assembly_planners)]
    releasable_assembly = releasable_results[releasable_results['Planner'].isin(assembly_planners)]
    total_assembly_count = len(total_assembly)
    total_assembly_hours = total_assembly['Hours'].sum()
    total_assembly_qty = total_assembly['Demand'].sum()
    releasable_assembly_count = len(releasable_assembly)
    releasable_assembly_hours = releasable_assembly['Hours'].sum()
    releasable_assembly_qty = releasable_assembly['Demand'].sum()
    
    # Packaging (Planner code 3804)
    packaging_planners = ['3804']
    total_packaging = df_results[df_results['Planner'].isin(packaging_planners)]
    releasable_packaging = releasable_results[releasable_results['Planner'].isin(packaging_planners)]
    total_packaging_count = len(total_packaging)
    total_packaging_hours = total_packaging['Hours'].sum()
    total_packaging_qty = total_packaging['Demand'].sum()
    releasable_packaging_count = len(releasable_packaging)
    releasable_packaging_hours = releasable_packaging['Hours'].sum()
    releasable_packaging_qty = releasable_packaging['Demand'].sum()
    
    # Malosa Instruments (Planner code 3805)
    malosa_instrument_planners = ['3805']
    total_malosa_instruments = df_results[df_results['Planner'].isin(malosa_instrument_planners)]
    releasable_malosa_instruments = releasable_results[releasable_results['Planner'].isin(malosa_instrument_planners)]
    total_malosa_instruments_count = len(total_malosa_instruments)
    total_malosa_instruments_hours = total_malosa_instruments['Hours'].sum()
    total_malosa_instruments_qty = total_malosa_instruments['Demand'].sum()
    releasable_malosa_instruments_count = len(releasable_malosa_instruments)
    releasable_malosa_instruments_hours = releasable_malosa_instruments['Hours'].sum()
    releasable_malosa_instruments_qty = releasable_malosa_instruments['Demand'].sum()
    
    # Virtuoso (Planner code 3806)
    virtuoso_planners = ['3806']
    total_virtuoso = df_results[df_results['Planner'].isin(virtuoso_planners)]
    releasable_virtuoso = releasable_results[releasable_results['Planner'].isin(virtuoso_planners)]
    total_virtuoso_count = len(total_virtuoso)
    total_virtuoso_hours = total_virtuoso['Hours'].sum()
    total_virtuoso_qty = total_virtuoso['Demand'].sum()
    releasable_virtuoso_count = len(releasable_virtuoso)
    releasable_virtuoso_hours = releasable_virtuoso['Hours'].sum()
    releasable_virtuoso_qty = releasable_virtuoso['Demand'].sum()
    
    # Kit Samples (Planner code KIT SAMPLES)
    kit_samples_planners = ['KIT SAMPLES']
    total_kit_samples = df_results[df_results['Planner'].isin(kit_samples_planners)]
    releasable_kit_samples = releasable_results[releasable_results['Planner'].isin(kit_samples_planners)]
    total_kit_samples_count = len(total_kit_samples)
    total_kit_samples_hours = total_kit_samples['Hours'].sum()
    total_kit_samples_qty = total_kit_samples['Demand'].sum()
    releasable_kit_samples_count = len(releasable_kit_samples)
    releasable_kit_samples_hours = releasable_kit_samples['Hours'].sum()
    releasable_kit_samples_qty = releasable_kit_samples['Demand'].sum()
    
    # Total Instruments
    total_instruments_count = (total_manufacturing_count + total_assembly_count + 
                             total_packaging_count + total_malosa_instruments_count)
    total_instruments_hours = (total_manufacturing_hours + total_assembly_hours + 
                             total_packaging_hours + total_malosa_instruments_hours)
    total_instruments_qty = (total_manufacturing_qty + total_assembly_qty + 
                           total_packaging_qty + total_malosa_instruments_qty)
    releasable_instruments_count = (releasable_manufacturing_count + releasable_assembly_count + 
                                  releasable_packaging_count + releasable_malosa_instruments_count)
    releasable_instruments_hours = (releasable_manufacturing_hours + releasable_assembly_hours + 
                                  releasable_packaging_hours + releasable_malosa_instruments_hours)
    releasable_instruments_qty = (releasable_manufacturing_qty + releasable_assembly_qty + 
                                releasable_packaging_qty + releasable_malosa_instruments_qty)
    
    # DEBUG: Show final results summary
    if DEBUG_MODE:
        print(f"\n🔍 DEBUG: Processing complete for {scenario_name}")
        print(f"   Total orders processed: {total_orders}")
        print(f"   Releasable orders: {releasable_count}")
        print(f"   Held orders: {held_count}")
        print(f"   Release rate: {releasable_count/total_orders*100:.1f}%" if total_orders > 0 else "   Release rate: 0%")
        print(f"   Total hours: {total_hours:.1f}")
        print(f"   Releasable hours: {releasable_hours:.1f}")
        print(f"   Hours release rate: {releasable_hours/total_hours*100:.1f}%" if total_hours > 0 else "   Hours release rate: 0%")
        
        # Show debug component summary if tracking specific part
        if DEBUG_COMPONENT_PART is not None:
            debug_part = str(DEBUG_COMPONENT_PART)
            initial_stock = stock.get(debug_part, 0)
            final_used = used_components.get(debug_part, 0)
            remaining_stock = initial_stock - final_used
            
            print(f"\n  🎯 DEBUG COMPONENT SUMMARY ({debug_part}):")
            print(f"    Orders using this component: {debug_component_orders_found}")
            print(f"    Initial stock:     {initial_stock}")
            print(f"    Total allocated:   {final_used}")
            print(f"    Remaining stock:   {remaining_stock}")
            print(f"    Allocation rate:   {final_used/initial_stock*100:.1f}%" if initial_stock > 0 else "    Allocation rate:   0%")
            
            # Show all Shop Orders that tried to allocate this component
            print(f"\n  📋 ALL SHOP ORDERS THAT TRIED TO ALLOCATE {debug_part}:")
            component_allocation_count = 0
            for result in results:
                result_so = result["SO Number"]
                result_part = result["Part"]
                result_status = result["Status"]
                
                # Check if this SO directly uses the component as parent part
                if str(result_part) == debug_part:
                    component_allocation_count += 1
                    print(f"    {component_allocation_count:2d}. SO {result_so}: Direct use as parent part - {result_status}")
                else:
                    # Check if this SO's BOM contains the component
                    try:
                        bom_check = planned_demand[planned_demand["SO Number"] == result_so]
                        if not bom_check.empty:
                            bom_components = bom_check["Component Part Number"].astype(str).tolist()
                            if debug_part in bom_components:
                                component_allocation_count += 1
                                comp_row = bom_check[bom_check["Component Part Number"].astype(str) == debug_part].iloc[0]
                                comp_qty = comp_row["Component Qty Required"]
                                print(f"    {component_allocation_count:2d}. SO {result_so} (Parent: {result_part}): Requires {comp_qty} units - {result_status}")
                    except:
                        pass
            
            if component_allocation_count == 0:
                print(f"    No Shop Orders found that use component {debug_part}")
            else:
                print(f"\n    Total Shop Orders trying to allocate {debug_part}: {component_allocation_count}")
        
        print("="*60)
    
    # End order processing and total processing phases
    performance_tracker.end_phase()  # End Order Processing
    performance_tracker.end_phase()  # End Total Processing
    
    return {
        'name': scenario_name,
        'filepath': 'Database',
        'sorting_strategy': sorting_strategy["name"] if sorting_strategy else "Default (Start Date)",
        'results_df': df_results,
        'metrics': {
            'total_orders': total_orders,
            'releasable_count': releasable_count,
            'held_count': held_count,
            'pb_count': pb_count,
            'skipped_count': skipped_count,
            '---1': '---',
            'total_hours': total_hours,
            'releasable_hours': releasable_hours,
            'held_hours': held_hours,
            'total_qty': total_qty,
            'releasable_qty': releasable_qty,
            'held_qty': held_qty,
            '---2': '---',
            'total_kits_count': total_kits_count,
            'total_kits_hours': total_kits_hours,
            'total_kits_qty': total_kits_qty,
            'releasable_kits_count': releasable_kits_count,
            'releasable_kits_hours': releasable_kits_hours,
            'releasable_kits_qty': releasable_kits_qty,
            'total_bvi_kits_count': total_bvi_kits_count,
            'total_bvi_kits_hours': total_bvi_kits_hours,
            'total_bvi_kits_qty': total_bvi_kits_qty,
            'releasable_bvi_kits_count': releasable_bvi_kits_count,
            'releasable_bvi_kits_hours': releasable_bvi_kits_hours,
            'releasable_bvi_kits_qty': releasable_bvi_kits_qty,
            'total_malosa_kits_count': total_malosa_kits_count,
            'total_malosa_kits_hours': total_malosa_kits_hours,
            'total_malosa_kits_qty': total_malosa_kits_qty,
            'releasable_malosa_kits_count': releasable_malosa_kits_count,
            'releasable_malosa_kits_hours': releasable_malosa_kits_hours,
            'releasable_malosa_kits_qty': releasable_malosa_kits_qty,
            '---3': '---',
            'total_instruments_count': total_instruments_count,
            'total_instruments_hours': total_instruments_hours,
            'total_instruments_qty': total_instruments_qty,
            'releasable_instruments_count': releasable_instruments_count,
            'releasable_instruments_hours': releasable_instruments_hours,
            'releasable_instruments_qty': releasable_instruments_qty,
            'total_manufacturing_count': total_manufacturing_count,
            'total_manufacturing_hours': total_manufacturing_hours,
            'total_manufacturing_qty': total_manufacturing_qty,
            'releasable_manufacturing_count': releasable_manufacturing_count,
            'releasable_manufacturing_hours': releasable_manufacturing_hours,
            'releasable_manufacturing_qty': releasable_manufacturing_qty,
            'total_assembly_count': total_assembly_count,
            'total_assembly_hours': total_assembly_hours,
            'total_assembly_qty': total_assembly_qty,
            'releasable_assembly_count': releasable_assembly_count,
            'releasable_assembly_hours': releasable_assembly_hours,
            'releasable_assembly_qty': releasable_assembly_qty,
            'total_packaging_count': total_packaging_count,
            'total_packaging_hours': total_packaging_hours,
            'total_packaging_qty': total_packaging_qty,
            'releasable_packaging_count': releasable_packaging_count,
            'releasable_packaging_hours': releasable_packaging_hours,
            'releasable_packaging_qty': releasable_packaging_qty,
            'total_malosa_instruments_count': total_malosa_instruments_count,
            'total_malosa_instruments_hours': total_malosa_instruments_hours,
            'total_malosa_instruments_qty': total_malosa_instruments_qty,
            'releasable_malosa_instruments_count': releasable_malosa_instruments_count,
            'releasable_malosa_instruments_hours': releasable_malosa_instruments_hours,
            'releasable_malosa_instruments_qty': releasable_malosa_instruments_qty,
            '---4': '---',
            'total_virtuoso_count': total_virtuoso_count,
            'total_virtuoso_hours': total_virtuoso_hours,
            'total_virtuoso_qty': total_virtuoso_qty,
            'releasable_virtuoso_count': releasable_virtuoso_count,
            'releasable_virtuoso_hours': releasable_virtuoso_hours,
            'releasable_virtuoso_qty': releasable_virtuoso_qty,
            '---5': '---',
            'committed_parts_count': committed_parts_count,
            'total_committed_qty': total_committed_qty,
            'total_kit_samples_count': total_kit_samples_count,
            'total_kit_samples_hours': total_kit_samples_hours,
            'total_kit_samples_qty': total_kit_samples_qty,
            'releasable_kit_samples_count': releasable_kit_samples_count,
            'releasable_kit_samples_hours': releasable_kit_samples_hours,
            'releasable_kit_samples_qty': releasable_kit_samples_qty
        }
    }

def load_and_process_database():
    global quick_analysis_excel_buffer
    """Load and process data from database instead of Excel files"""
    
    # Reset performance tracker for this run
    performance_tracker.cleanup()
    
    # Test database connection first
    connection_success, connection_message = test_database_connection()
    if not connection_success:
        messagebox.showerror("Database Connection Error", connection_message)
        status_var.set("❌ Database connection failed")
        return

    try:
        # Clear the UI and start fresh
        status_var.set("🔄 Initializing database connection...")
        root.update_idletasks()
        
        start_time = time.time()
        main_frame.configure(style='TFrame')
        
        scenarios = []
        scenarios_for_comparison = []  # Will store all tested scenarios for comparison tables
        
        # Progress callback function that updates UI immediately
        def update_progress(message):
            status_var.set(message)
            root.update_idletasks()
        
        # Determine processing mode
        minmax_mode = minmax_var.get()
        
        if minmax_mode:
            # Min/Max optimization mode - test all sorting strategies
            strategies = get_sorting_strategies()
            total_scenarios = len(strategies)
            
            update_progress(f"🔥 MIN/MAX MODE: Testing {len(strategies)} sorting strategies on database = {total_scenarios} total scenarios")
            time.sleep(1)
            
            scenario_num = 0
            all_strategy_results = []  # Store ALL results for comparison
            
            for strategy_idx, strategy in enumerate(strategies):
                scenario_num += 1
                scenario_name = f"Database_{strategy['name'].replace(' ', '_').replace('(', '').replace(')', '')}"
                
                # Process with specific sorting strategy
                scenario_start_time = time.time()
                scenario_result = process_single_scenario(
                    scenario_name, update_progress, 
                    scenario_num, total_scenarios, strategy,
                    include_kits=include_kits_var.get(),
                    include_instruments=include_instruments_var.get(),
                    include_virtuoso=include_virtuoso_var.get(),
                    include_kit_samples=include_kit_samples_var.get()
                )
                scenario_end_time = time.time()
                scenario_duration = scenario_end_time - scenario_start_time
                
                # Store result for this strategy
                all_strategy_results.append(scenario_result)
                
                # Show completion
                metrics = scenario_result['metrics']
                remaining_scenarios = total_scenarios - scenario_num
                
                if remaining_scenarios > 0:
                    total_elapsed = time.time() - start_time
                    avg_time_per_scenario = total_elapsed / scenario_num
                    estimated_remaining = remaining_scenarios * avg_time_per_scenario
                    
                    update_progress(f"✅ [{scenario_num}/{total_scenarios}] {strategy['name']}: {metrics['releasable_count']:,}/{metrics['total_orders']:,} orders ({scenario_duration:.1f}s) | {estimated_remaining:.0f}s remaining")
                else:
                    update_progress(f"✅ [{scenario_num}/{total_scenarios}] {strategy['name']}: {metrics['releasable_count']:,}/{metrics['total_orders']:,} orders ({scenario_duration:.1f}s) | OPTIMIZATION COMPLETE!")
                
                time.sleep(0.2)  # Brief pause between strategies
            
            # Find the best strategies
            best_orders_strategy = max(all_strategy_results, key=lambda s: s['metrics']['releasable_count'])
            best_hours_strategy = max(all_strategy_results, key=lambda s: s['metrics']['releasable_hours'])
            best_qty_strategy = max(all_strategy_results, key=lambda s: s['metrics']['releasable_qty'])
            
            # Create NEW scenario objects with clear names for the best strategies
            # Best Orders Strategy
            best_orders_scenario = {
                'name': f"BEST_ORDERS_Database",
                'filepath': 'Database',
                'sorting_strategy': f"🏆 BEST ORDERS: {best_orders_strategy['sorting_strategy']}",
                'results_df': best_orders_strategy['results_df'],
                'metrics': best_orders_strategy['metrics']
            }
            scenarios.append(best_orders_scenario)
            
            # Best Hours Strategy
            best_hours_scenario = {
                'name': f"BEST_HOURS_Database",
                'filepath': 'Database',
                'sorting_strategy': f"🏆 BEST HOURS: {best_hours_strategy['sorting_strategy']}",
                'results_df': best_hours_strategy['results_df'],
                'metrics': best_hours_strategy['metrics']
            }
            scenarios.append(best_hours_scenario)
            
            # Best Quantity Strategy
            best_qty_scenario = {
                'name': f"BEST_QTY_Database",
                'filepath': 'Database',
                'sorting_strategy': f"🏆 BEST QTY: {best_qty_strategy['sorting_strategy']}",
                'results_df': best_qty_strategy['results_df'],
                'metrics': best_qty_strategy['metrics']
            }
            scenarios.append(best_qty_scenario)
            
            update_progress(f"🏆 Database optimized: Orders={best_orders_strategy['sorting_strategy']} ({best_orders_strategy['metrics']['releasable_count']:,}), Hours={best_hours_strategy['sorting_strategy']} ({best_hours_strategy['metrics']['releasable_hours']:,.0f}), Qty={best_qty_strategy['sorting_strategy']} ({best_qty_strategy['metrics']['releasable_qty']:,})")
            time.sleep(0.5)
            
            # Use all_strategy_results for comparison tables
            scenarios_for_comparison = all_strategy_results
        else:
            # Standard mode - process database once
            total_scenarios = 1
            
            scenario_name = "Database_Scenario"
            scenario_num = 1
            
            update_progress(f"📊 [Scenario {scenario_num}/{total_scenarios}] Starting: Database Analysis")
            
            # Process with live progress updates
            scenario_start_time = time.time()
            scenario_result = process_single_scenario(
                scenario_name, update_progress, scenario_num, total_scenarios,
                include_kits=include_kits_var.get(),
                include_instruments=include_instruments_var.get(),
                include_virtuoso=include_virtuoso_var.get(),
                include_kit_samples=include_kit_samples_var.get()
            )
            scenario_end_time = time.time()
            scenario_duration = scenario_end_time - scenario_start_time
            
            scenarios.append(scenario_result)
            scenarios_for_comparison.append(scenario_result)  # Same as scenarios in standard mode
            
            # Show completion with actual metrics and time
            metrics = scenario_result['metrics']
            update_progress(f"✅ [Scenario {scenario_num}/{total_scenarios}] Complete: {metrics['releasable_count']:,}/{metrics['total_orders']:,} releasable ({scenario_duration:.1f}s) | COMPLETE!")
            
            time.sleep(0.3)
        
        # Calculate total processing time
        end_time = time.time()
        processing_time = end_time - start_time
        
        # Helper function to aggregate metrics from a list of scenarios
        def aggregate_metrics(scenario_list, is_min_max_mode_for_totals=False):
            if not scenario_list:
                # Return a default structure if the list is empty
                return {key: 0 for key in [
                    'total_qty', 'total_hours',
                    'releasable_count', 'releasable_qty', 'releasable_hours',
                    'releasable_kits_count', 'releasable_kits_hours', 'releasable_kits_qty',
                    'releasable_bvi_kits_count', 'releasable_bvi_kits_hours', 'releasable_bvi_kits_qty',
                    'releasable_malosa_kits_count', 'releasable_malosa_kits_hours', 'releasable_malosa_kits_qty',
                    'releasable_instruments_count', 'releasable_instruments_hours', 'releasable_instruments_qty',
                    'releasable_manufacturing_count', 'releasable_manufacturing_hours', 'releasable_manufacturing_qty',
                    'releasable_assembly_count', 'releasable_assembly_hours', 'releasable_assembly_qty',
                    'releasable_packaging_count', 'releasable_packaging_hours', 'releasable_packaging_qty',
                    'releasable_malosa_instruments_count', 'releasable_malosa_instruments_hours', 'releasable_malosa_instruments_qty',
                    'releasable_virtuoso_count', 'releasable_virtuoso_hours', 'releasable_virtuoso_qty',
                    'total_qty', 'total_hours' # These are per-scenario totals, summing them gives overall totals
                ]}

            agg = {}
            keys_to_sum = [
                'releasable_count', 'releasable_qty', 'releasable_hours',
                'releasable_kits_count', 'releasable_kits_hours', 'releasable_kits_qty',
                'releasable_bvi_kits_count', 'releasable_bvi_kits_hours', 'releasable_bvi_kits_qty',
                'releasable_malosa_kits_count', 'releasable_malosa_kits_hours', 'releasable_malosa_kits_qty',
                'releasable_instruments_count', 'releasable_instruments_hours', 'releasable_instruments_qty',
                'releasable_manufacturing_count', 'releasable_manufacturing_hours', 'releasable_manufacturing_qty',
                'releasable_assembly_count', 'releasable_assembly_hours', 'releasable_assembly_qty',
                'releasable_packaging_count', 'releasable_packaging_hours', 'releasable_packaging_qty',
                'releasable_malosa_instruments_count', 'releasable_malosa_instruments_hours', 'releasable_malosa_instruments_qty',
                'releasable_virtuoso_count', 'releasable_virtuoso_hours', 'releasable_virtuoso_qty',
                'total_qty', 'total_hours' # These are per-scenario totals, summing them gives overall totals
            ]
            for key in keys_to_sum:
                agg[key] = sum(s['metrics'].get(key, 0) for s in scenario_list)
            
            return agg

        # Determine source metrics for the Summary sheet
        source_metrics = {}
        total_orders_processed = 0

        if not scenarios: # Handle case where no scenarios were processed (e.g., database connection failed)
            # Use default empty metrics for the summary if no scenarios exist
            source_metrics = aggregate_metrics([], is_min_max_mode_for_totals=minmax_mode)
            # total_orders_processed is already 0
            # Ensure we have a default for "Files Processed"
            files_processed_list_for_summary = []
        else:
            files_processed_list_for_summary = ['Database'] # Use database as the source

            if minmax_mode:
                # In min/max mode, total_orders_processed sums total_orders from one "best" strategy per unique file
                seen_files_for_total_orders = set()
                unique_scenarios_for_total_orders = []
                for s in scenarios: # 'scenarios' contains BEST_ORDERS, BEST_HOURS, BEST_QTY per file
                    if s['filepath'] not in seen_files_for_total_orders:
                        unique_scenarios_for_total_orders.append(s)
                        seen_files_for_total_orders.add(s['filepath'])
                total_orders_processed = sum(s['metrics']['total_orders'] for s in unique_scenarios_for_total_orders)
                
                # For detailed metrics in summary, aggregate from "BEST_ORDERS" strategies
                best_orders_scenarios = [s for s in scenarios if "BEST_ORDERS" in s['name']]
                source_metrics = aggregate_metrics(best_orders_scenarios, is_min_max_mode_for_totals=True)
            else: # Standard mode
                total_orders_processed = sum(s['metrics']['total_orders'] for s in scenarios)
                if len(scenarios) == 1:
                    source_metrics = scenarios[0]['metrics']
                else: # Standard mode, multiple files
                    source_metrics = aggregate_metrics(scenarios)
        
        orders_per_second = total_orders_processed / processing_time if processing_time > 0 else 0
        
        # Create summary sheet (move this above export logic so it's always defined)
        summary_items = [
            ('Tool Version', f"{VERSION} ({VERSION_DATE})"),
            ('Processing Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Processing Mode', "Min/Max Optimization" if minmax_mode else "Standard"),
            ('Files Processed Count', len(files_processed_list_for_summary)),
            ('Strategies Tested / Scenarios', len(scenarios_for_comparison) if minmax_mode else len(scenarios)),
            ('Optimal Strategies / Scenarios Saved', len(scenarios)),
            ('--- Material Categories Processed ---', '---'),
            ('Kits Included', "Yes" if include_kits_var.get() else "No"),
            ('Instruments Included', "Yes" if include_instruments_var.get() else "No"),
            ('Virtuoso Included', "Yes" if include_virtuoso_var.get() else "No"),
            ('Kit Samples Included', "Yes" if include_kit_samples_var.get() else "No"),
            ('--- Overall Performance ---', '---'),
            ('Total Orders Processed', f"{total_orders_processed:,}"),
            ('Total Demand Quantity', f"{source_metrics.get('total_qty', 0):,}"),
            ('Total Demand Hours', f"{source_metrics.get('total_hours', 0):,.1f}"),
            ('Releasable Orders', f"{source_metrics.get('releasable_count', 0):,}"),
            ('Releasable Quantity', f"{source_metrics.get('releasable_qty', 0):,.1f}"),
            ('Releasable Hours', f"{source_metrics.get('releasable_hours', 0):,.1f}"),
            ('--- Total Kits Processed ---', '---'),
            ('Total Kits Processed (Orders)', f"{source_metrics.get('total_kits_count', 0):,}"),
            ('Total Kits Processed (Hours)', f"{source_metrics.get('total_kits_hours', 0):,.1f}"),
            ('Total Kits Processed (Quantity)', f"{source_metrics.get('total_kits_qty', 0):,}"),
            ('--- Releasable Kits Breakdown ---', '---'),
            ('Total Releasable Kits (Orders)', f"{source_metrics.get('releasable_kits_count', 0):,}"),
            ('Total Releasable Kits (Hours)', f"{source_metrics.get('releasable_kits_hours', 0):,.1f}"),
            ('Total Releasable Kits (Quantity)', f"{source_metrics.get('releasable_kits_qty', 0):,}"),
            ('Kits Release Rate (Orders %)', f"{source_metrics.get('releasable_kits_count', 0)/source_metrics.get('total_kits_count', 1)*100:.1f}%" if source_metrics.get('total_kits_count', 0) > 0 else "0%"),
            ('Kits Release Rate (Hours %)', f"{source_metrics.get('releasable_kits_hours', 0)/source_metrics.get('total_kits_hours', 1)*100:.1f}%" if source_metrics.get('total_kits_hours', 0) > 0 else "0%"),
            ('Kits Release Rate (Quantity %)', f"{source_metrics.get('releasable_kits_qty', 0)/source_metrics.get('total_kits_qty', 1)*100:.1f}%" if source_metrics.get('total_kits_qty', 0) > 0 else "0%"),
            ('  BVI Kits (Orders)', f"{source_metrics.get('releasable_bvi_kits_count', 0):,}"),
            ('  BVI Kits (Hours)', f"{source_metrics.get('releasable_bvi_kits_hours', 0):,.1f}"),
            ('  BVI Kits (Quantity)', f"{source_metrics.get('releasable_bvi_kits_qty', 0):,}"),
            ('  BVI Kits Release Rate (%)', f"{source_metrics.get('releasable_bvi_kits_count', 0)/source_metrics.get('total_bvi_kits_count', 1)*100:.1f}%" if source_metrics.get('total_bvi_kits_count', 0) > 0 else "0%"),
            ('  Malosa Kits (Orders)', f"{source_metrics.get('releasable_malosa_kits_count', 0):,}"),
            ('  Malosa Kits (Hours)', f"{source_metrics.get('releasable_malosa_kits_hours', 0):,.1f}"),
            ('  Malosa Kits (Quantity)', f"{source_metrics.get('releasable_malosa_kits_qty', 0):,}"),
            ('  Malosa Kits Release Rate (%)', f"{source_metrics.get('releasable_malosa_kits_count', 0)/source_metrics.get('total_malosa_kits_count', 1)*100:.1f}%" if source_metrics.get('total_malosa_kits_count', 0) > 0 else "0%"),
            ('--- Total Instruments Processed ---', '---'),
            ('Total Instruments Processed (Orders)', f"{source_metrics.get('total_instruments_count', 0):,}"),
            ('Total Instruments Processed (Hours)', f"{source_metrics.get('total_instruments_hours', 0):,.1f}"),
            ('Total Instruments Processed (Quantity)', f"{source_metrics.get('total_instruments_qty', 0):,}"),
            ('--- Releasable Instruments Breakdown ---', '---'),
            ('Total Releasable Instruments (Orders)', f"{source_metrics.get('releasable_instruments_count', 0):,}"),
            ('Total Releasable Instruments (Hours)', f"{source_metrics.get('releasable_instruments_hours', 0):,.1f}"),
            ('Total Releasable Instruments (Quantity)', f"{source_metrics.get('releasable_instruments_qty', 0):,}"),
            ('Instruments Release Rate (Orders %)', f"{source_metrics.get('releasable_instruments_count', 0)/source_metrics.get('total_instruments_count', 1)*100:.1f}%" if source_metrics.get('total_instruments_count', 0) > 0 else "0%"),
            ('Instruments Release Rate (Hours %)', f"{source_metrics.get('releasable_instruments_hours', 0)/source_metrics.get('total_instruments_hours', 1)*100:.1f}%" if source_metrics.get('total_instruments_hours', 0) > 0 else "0%"),
            ('Instruments Release Rate (Quantity %)', f"{source_metrics.get('releasable_instruments_qty', 0)/source_metrics.get('total_instruments_qty', 1)*100:.1f}%" if source_metrics.get('total_instruments_qty', 0) > 0 else "0%"),
            ('  Manufacturing (3802 Orders)', f"{source_metrics.get('releasable_manufacturing_count', 0):,}"),
            ('  Manufacturing (3802 Hours)', f"{source_metrics.get('releasable_manufacturing_hours', 0):,.1f}"),
            ('  Manufacturing (3802 Quantity)', f"{source_metrics.get('releasable_manufacturing_qty', 0):,}"),
            ('  Manufacturing Release Rate (%)', f"{source_metrics.get('releasable_manufacturing_count', 0)/source_metrics.get('total_manufacturing_count', 1)*100:.1f}%" if source_metrics.get('total_manufacturing_count', 0) > 0 else "0%"),
            ('  Assembly (3803 Orders)', f"{source_metrics.get('releasable_assembly_count', 0):,}"),
            ('  Assembly (3803 Hours)', f"{source_metrics.get('releasable_assembly_hours', 0):,.1f}"),
            ('  Assembly (3803 Quantity)', f"{source_metrics.get('releasable_assembly_qty', 0):,}"),
            ('  Assembly Release Rate (%)', f"{source_metrics.get('releasable_assembly_count', 0)/source_metrics.get('total_assembly_count', 1)*100:.1f}%" if source_metrics.get('total_assembly_count', 0) > 0 else "0%"),
            ('  Packaging (3804 Orders)', f"{source_metrics.get('releasable_packaging_count', 0):,}"),
            ('  Packaging (3804 Hours)', f"{source_metrics.get('releasable_packaging_hours', 0):,.1f}"),
            ('  Packaging (3804 Quantity)', f"{source_metrics.get('releasable_packaging_qty', 0):,}"),
            ('  Packaging Release Rate (%)', f"{source_metrics.get('releasable_packaging_count', 0)/source_metrics.get('total_packaging_count', 1)*100:.1f}%" if source_metrics.get('total_packaging_count', 0) > 0 else "0%"),
            ('  Malosa Instruments (3805 Orders)', f"{source_metrics.get('releasable_malosa_instruments_count', 0):,}"),
            ('  Malosa Instruments (3805 Hours)', f"{source_metrics.get('releasable_malosa_instruments_hours', 0):,.1f}"),
            ('  Malosa Instruments (3805 Quantity)', f"{source_metrics.get('releasable_malosa_instruments_qty', 0):,}"),
            ('  Malosa Instruments Release Rate (%)', f"{source_metrics.get('releasable_malosa_instruments_count', 0)/source_metrics.get('total_malosa_instruments_count', 1)*100:.1f}%" if source_metrics.get('total_malosa_instruments_count', 0) > 0 else "0%"),
            ('--- Total Virtuoso Processed ---', '---'),
            ('Total Virtuoso Processed (Orders)', f"{source_metrics.get('total_virtuoso_count', 0):,}"),
            ('Total Virtuoso Processed (Hours)', f"{source_metrics.get('total_virtuoso_hours', 0):,.1f}"),
            ('Total Virtuoso Processed (Quantity)', f"{source_metrics.get('total_virtuoso_qty', 0):,}"),
            ('--- Releasable Virtuoso Breakdown ---', '---'),
            ('Releasable Virtuoso (Orders)', f"{source_metrics.get('releasable_virtuoso_count', 0):,}"),
            ('Releasable Virtuoso (Hours)', f"{source_metrics.get('releasable_virtuoso_hours', 0):,.1f}"),
            ('Releasable Virtuoso (Quantity)', f"{source_metrics.get('releasable_virtuoso_qty', 0):,}"),
            ('Virtuoso Release Rate (Orders %)', f"{source_metrics.get('releasable_virtuoso_count', 0)/source_metrics.get('total_virtuoso_count', 1)*100:.1f}%" if source_metrics.get('total_virtuoso_count', 0) > 0 else "0%"),
            ('Virtuoso Release Rate (Hours %)', f"{source_metrics.get('releasable_virtuoso_hours', 0)/source_metrics.get('total_virtuoso_hours', 1)*100:.1f}%" if source_metrics.get('total_virtuoso_hours', 0) > 0 else "0%"),
            ('Virtuoso Release Rate (Quantity %)', f"{source_metrics.get('releasable_virtuoso_qty', 0)/source_metrics.get('total_virtuoso_qty', 1)*100:.1f}%" if source_metrics.get('total_virtuoso_qty', 0) > 0 else "0%"),
            ('--- Processing Performance ---', '---'),
            ('Total Processing Time (seconds)', f"{processing_time:.2f}"),
            ('Processing Speed (orders/second)', f"{orders_per_second:.1f}"),
            ('Avg Time per Strategy/Scenario (seconds)', f"{processing_time/len(scenarios_for_comparison):.2f}" if minmax_mode and scenarios_for_comparison else (f"{processing_time/len(scenarios):.2f}" if scenarios else "N/A")),
            ('Processed File Names', "; ".join([os.path.basename(f) for f in files_processed_list_for_summary]))
        ]
        summary_data = pd.DataFrame(summary_items, columns=['Metric', 'Value'])
        
        # Save results
        if not no_export_var.get():
            quick_analysis_excel_buffer = None  # Not used in normal mode
            # Save results directly to desktop
            desktop_path = os.path.join(os.path.expanduser("~"), "OneDrive - BVI\Desktop")
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            
            if minmax_mode:
                output_file = os.path.join(desktop_path, f"MinMax_Optimization_Analysis_{VERSION}_{timestamp}.xlsx")
            elif len(scenarios) > 1:
                output_file = os.path.join(desktop_path, f"Multi_Scenario_Analysis_{VERSION}_{timestamp}.xlsx")
            else:
                output_file = os.path.join(desktop_path, f"Material_Release_Plan_{VERSION}_{timestamp}.xlsx")
            
            status_var.set("💾 Saving optimization results...")
            root.update_idletasks()
            
            # Create comparison data if multiple scenarios
            if len(scenarios_for_comparison) > 1:
                comparison_data = []
                for scenario in scenarios_for_comparison:
                    metrics = scenario['metrics']
                    comparison_data.append({
                        'Scenario': scenario['name'],
                        'File': os.path.basename(scenario['filepath']),
                        'Sorting Strategy': scenario['sorting_strategy'],
                        'Total Orders': metrics['total_orders'],
                        'Releasable Orders': metrics['releasable_count'],
                        'Held Orders': metrics['held_count'],
                        'Release Rate (%)': f"{metrics['releasable_count']/metrics['total_orders']*100:.1f}%" if metrics['total_orders'] > 0 else "0%",
                        '---1': '---',
                        'Total Qty': f"{metrics['total_qty']:,}",
                        'Releasable Qty': f"{metrics['releasable_qty']:,}",
                        'Qty Release Rate (%)': f"{metrics['releasable_qty']/metrics['total_qty']*100:.1f}%" if metrics['total_qty'] > 0 else "0%",
                        'Piggyback Orders': metrics['pb_count'],
                        'Total Hours': f"{metrics['total_hours']:,.1f}",
                        'Releasable Hours': f"{metrics['releasable_hours']:,.1f}",
                        'Labor Release Rate (%)': f"{metrics['releasable_hours']/metrics['total_hours']*100:.1f}%" if metrics['total_hours'] > 0 else "0%",
                        '---2': '---',
                        'BVI Kits': metrics['releasable_bvi_kits_count'],
                        'BVI Kit Hours': f"{metrics['releasable_bvi_kits_hours']:,.1f}",
                        'BVI Kit Qty': f"{metrics['releasable_bvi_kits_qty']:,}",
                        'Malosa Kits': metrics['releasable_malosa_kits_count'],
                        'Malosa Kit Hours': f"{metrics['releasable_malosa_kits_hours']:,.1f}",
                        'Malosa Kit Qty': f"{metrics['releasable_malosa_kits_qty']:,}",
                        'Total Kits': metrics['releasable_kits_count'],
                        'Total Kit Hours': f"{metrics['releasable_kits_hours']:,.1f}",
                        'Total Kit Qty': f"{metrics['releasable_kits_qty']:,}",
                        '---3': '---',
                        'Manufacturing (3802)': metrics['releasable_manufacturing_count'],
                        'Manufacturing Hours': f"{metrics['releasable_manufacturing_hours']:,.1f}",
                        'Manufacturing Qty': f"{metrics['releasable_manufacturing_qty']:,}",
                        'Assembly (3803)': metrics['releasable_assembly_count'],
                        'Assembly Hours': f"{metrics['releasable_assembly_hours']:,.1f}",
                        'Assembly Qty': f"{metrics['releasable_assembly_qty']:,}",
                        'Packaging (3804)': metrics['releasable_packaging_count'],
                        'Packaging Hours': f"{metrics['releasable_packaging_hours']:,.1f}",
                        'Packaging Qty': f"{metrics['releasable_packaging_qty']:,}",
                        'Malosa Inst (3805)': metrics['releasable_malosa_instruments_count'],
                        'Malosa Inst Hours': f"{metrics['releasable_malosa_instruments_hours']:,.1f}",
                        'Malosa Inst Qty': f"{metrics['releasable_malosa_instruments_qty']:,}",
                        'Virtuoso (3806)': metrics['releasable_virtuoso_count'],
                        'Virtuoso Hours': f"{metrics['releasable_virtuoso_hours']:,.1f}",
                        'Virtuoso Qty': f"{metrics['releasable_virtuoso_qty']:,}",
                        'Total Instruments': metrics['releasable_instruments_count'],
                        'Total Inst Hours': f"{metrics['releasable_instruments_hours']:,.1f}",
                        'Total Inst Qty': f"{metrics['releasable_instruments_qty']:,}",
                        '---4': '---',
                        'Committed Parts': metrics['committed_parts_count'],
                        'Committed Qty': f"{metrics['total_committed_qty']:,}",
                        'Kit Samples': metrics['releasable_kit_samples_count'],
                        'Kit Samples Hours': f"{metrics['releasable_kit_samples_hours']:,.1f}",
                        'Kit Samples Qty': f"{metrics['releasable_kit_samples_qty']:,}"
                    })
                comparison_df = pd.DataFrame(comparison_data)

            # Write everything to Excel in a single writer session
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Define styles once at the start
                header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                separator_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                bold_font = Font(bold=True)

                # Write each scenario to its own sheet first
                for scenario in scenarios:
                    sheet_name = scenario['name'][:31]  # Excel sheet name limit
                    df = scenario['results_df'].copy()
                    
                    # Convert numeric columns to proper number format
                    numeric_columns = ['Demand', 'Hours']
                    for col in numeric_columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce')
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Get the worksheet and apply number formats
                    worksheet = writer.sheets[sheet_name]
                    for idx, col in enumerate(df.columns, 1):
                        col_letter = get_column_letter(idx)
                        if col == 'Hours':
                            # Format hours with 1 decimal place
                            for cell in worksheet[col_letter][1:]:
                                cell.number_format = '#,##0.0'
                        elif col == 'Demand':
                            # Format demand as whole numbers
                            for cell in worksheet[col_letter][1:]:
                                cell.number_format = '#,##0'
                
                # Write the summary sheet with formatting
                summary_data.to_excel(writer, sheet_name='Summary', index=False)
                worksheet = writer.sheets['Summary']
                
                # Apply number formats to summary sheet
                for row in range(2, worksheet.max_row + 1):
                    value_cell = worksheet.cell(row=row, column=2)
                    metric_cell = worksheet.cell(row=row, column=1)
                    
                    if any(term in metric_cell.value for term in ['Hours', 'Time']):
                        value_cell.number_format = '#,##0.0'
                    elif any(term in metric_cell.value for term in ['Orders', 'Count', 'Quantity']):
                        value_cell.number_format = '#,##0'
                    elif 'Rate' in metric_cell.value or 'Speed' in metric_cell.value:
                        value_cell.number_format = '#,##0.0'
                    
                    # Apply visual formatting to Summary sheet
                    if metric_cell.value.startswith('---'):
                        # Apply separator formatting
                        for col in range(1, 3):  # Columns A and B
                            cell = worksheet.cell(row=row, column=col)
                            cell.fill = separator_fill
                    elif any(metric_cell.value.startswith(prefix) for prefix in ['Total', 'Releasable', 'BVI', 'Malosa', 'Manufacturing', 'Assembly', 'Packaging', 'Virtuoso']):
                        # Bold important metrics
                        metric_cell.font = bold_font
                
                # Format header row in Summary
                for col in range(1, 3):  # Columns A and B
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = bold_font
                    cell.fill = header_fill
                
                # Auto-adjust column widths in Summary
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # Write comparison sheet if it exists
                if len(scenarios_for_comparison) > 1:
                    comparison_df_formatted = comparison_df.copy()
                    
                    # Convert string numbers back to numeric format
                    numeric_columns = [
                        'Total Orders', 'Releasable Orders', 'Held Orders', 'Piggyback Orders',
                        'Total Hours', 'Releasable Hours', 'BVI Kits', 'BVI Kit Hours', 'BVI Kit Qty',
                        'Malosa Kits', 'Malosa Kit Hours', 'Malosa Kit Qty', 'Total Kits', 'Total Kit Hours',
                        'Total Kit Qty', 'Manufacturing (3802)', 'Manufacturing Hours', 'Manufacturing Qty',
                        'Assembly (3803)', 'Assembly Hours', 'Assembly Qty', 'Packaging (3804)', 'Packaging Hours',
                        'Packaging Qty', 'Malosa Inst (3805)', 'Malosa Inst Hours', 'Malosa Inst Qty',
                        'Virtuoso (3806)', 'Virtuoso Hours', 'Virtuoso Qty', 'Total Instruments',
                        'Total Inst Hours', 'Total Inst Qty', 'Committed Parts', 'Committed Qty',
                        'Kit Samples', 'Kit Samples Hours', 'Kit Samples Qty'
                    ]
                    
                    for col in numeric_columns:
                        if col in comparison_df_formatted.columns:
                            # Remove commas and convert to numeric
                            comparison_df_formatted[col] = comparison_df_formatted[col].astype(str).str.replace(',', '').str.replace('$', '')
                            comparison_df_formatted[col] = pd.to_numeric(comparison_df_formatted[col], errors='coerce')
                    
                    # Handle percentage columns separately
                    pct_columns = ['Release Rate (%)', 'Qty Release Rate (%)', 'Labor Release Rate (%)']
                    for col in pct_columns:
                        if col in comparison_df_formatted.columns:
                            # Remove % sign and convert to numeric percentage
                            comparison_df_formatted[col] = comparison_df_formatted[col].astype(str).str.rstrip('%').astype(float) / 100
                    
                    comparison_df_formatted.to_excel(writer, sheet_name='Strategy Comparison', index=False)
                    
                    # Format the comparison sheet
                    comp_worksheet = writer.sheets['Strategy Comparison']
                    
                    # Apply number formats to all cells in numeric columns
                    for col_idx, col_name in enumerate(comparison_df_formatted.columns, 1):
                        col_letter = get_column_letter(col_idx)
                        
                        if col_name in pct_columns:
                            # Format as percentage
                            for cell in comp_worksheet[col_letter][1:]:
                                cell.number_format = '0.0%'
                        elif 'Hours' in str(col_name):
                            # Format with 1 decimal place
                            for cell in comp_worksheet[col_letter][1:]:
                                cell.number_format = '#,##0.0'
                        elif any(term in str(col_name) for term in ['Orders', 'Qty', 'Count', 'Parts']):
                            # Format as whole number with thousands separator
                            for cell in comp_worksheet[col_letter][1:]:
                                cell.number_format = '#,##0'
                    
                    # Format headers and apply column widths
                    for col in range(1, comp_worksheet.max_column + 1):
                        cell = comp_worksheet.cell(row=1, column=col)
                        cell.font = bold_font
                        cell.fill = header_fill
                        
                        # Auto-adjust column width
                        max_length = 0
                        column = [cell for cell in comp_worksheet[get_column_letter(col)]]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                            adjusted_width = min((max_length + 2), 50)  # Cap width at 50
                            comp_worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width
        else:
            output_file = None  # No file created

        # Display results
        if minmax_mode:
            # Min/Max optimization summary
            files_processed = list(set([s['filepath'] for s in scenarios_for_comparison]))
            
            summary_text = f"""🔥 MIN/MAX OPTIMIZATION COMPLETE!

📊 OPTIMIZATION ANALYSIS:
   Files Analyzed: {len(files_processed)}
   Sorting Strategies Tested: {len(get_sorting_strategies())}
   Total Strategy Tests: {len(scenarios_for_comparison)}
   Best Strategies Saved: {len(scenarios)} individual sheets (3 per file: Orders, Hours, Qty)

🔧 MATERIAL CATEGORIES PROCESSED:
   Kits (3001, 3801, 5001): {'✓ Included' if include_kits_var.get() else '✗ Excluded'}
   Instruments (3802, 3803, 3804, 3805): {'✓ Included' if include_instruments_var.get() else '✗ Excluded'}
   Virtuoso (3806): {'✓ Included' if include_virtuoso_var.get() else '✗ Excluded'}
   Kit Samples (KIT SAMPLES): {'✓ Included' if include_kit_samples_var.get() else '✗ Excluded'}

"""
            
            for filepath in files_processed:
                file_scenarios = [s for s in scenarios_for_comparison if s['filepath'] == filepath]
                best_orders = max(file_scenarios, key=lambda s: s['metrics']['releasable_count'])
                best_hours = max(file_scenarios, key=lambda s: s['metrics']['releasable_hours'])
                best_qty = max(file_scenarios, key=lambda s: s['metrics']['releasable_qty'])
                worst_orders = min(file_scenarios, key=lambda s: s['metrics']['releasable_count'])
                
                improvement_orders = best_orders['metrics']['releasable_count'] - worst_orders['metrics']['releasable_count']
                improvement_pct = improvement_orders / worst_orders['metrics']['total_orders'] * 100
                
                summary_text += f"""📁 FILE: {os.path.basename(filepath)}
   🏆 BEST STRATEGY (Orders): {best_orders['sorting_strategy']}
      → {best_orders['metrics']['releasable_count']:>6}/{best_orders['metrics']['total_orders']:>6} orders releasable ({best_orders['metrics']['releasable_count']/best_orders['metrics']['total_orders']*100:.1f}%)
      
      🔧 KITS:
        BVI Kits:          {best_orders['metrics']['releasable_bvi_kits_count']:>6} orders,  {best_orders['metrics']['releasable_bvi_kits_hours']:>8.1f} hrs,  {best_orders['metrics']['releasable_bvi_kits_qty']:>8} qty
        Malosa Kits:       {best_orders['metrics']['releasable_malosa_kits_count']:>6} orders,  {best_orders['metrics']['releasable_malosa_kits_hours']:>8.1f} hrs,  {best_orders['metrics']['releasable_malosa_kits_qty']:>8} qty
              → {best_orders['metrics']['releasable_count']:>6}/{best_orders['metrics']['total_orders']:>6} orders releasable ({best_orders['metrics']['releasable_count']/best_orders['metrics']['total_orders']*100:.1f}%)
        
        🔧 KITS:
        BVI Kits:          {best_orders['metrics']['releasable_bvi_kits_count']:>6} orders,  {best_orders['metrics']['releasable_bvi_kits_hours']:>8.1f} hrs,  {best_orders['metrics']['releasable_bvi_kits_qty']:>8} qty
        Malosa Kits:       {best_orders['metrics']['releasable_malosa_kits_count']:>6} orders,  {best_orders['metrics']['releasable_malosa_kits_hours']:>8.1f} hrs,  {best_orders['metrics']['releasable_malosa_kits_qty']:>8} qty
        Total Kits:        {best_orders['metrics']['releasable_kits_count']:>6} orders,  {best_orders['metrics']['releasable_kits_hours']:>8.1f} hrs,  {best_orders['metrics']['releasable_kits_qty']:>8} qty
        
        🔬 INSTRUMENTS:
        Manufacturing:     {best_orders['metrics']['releasable_manufacturing_count']:>6} orders,  {best_orders['metrics']['releasable_manufacturing_hours']:>8.1f} hrs,  {best_orders['metrics']['releasable_manufacturing_qty']:>8} qty
        Assembly:         {best_orders['metrics']['releasable_assembly_count']:>6} orders,  {best_orders['metrics']['releasable_assembly_hours']:>8.1f} hrs,  {best_orders['metrics']['releasable_assembly_qty']:>8} qty
        Packaging:        {best_orders['metrics']['releasable_packaging_count']:>6} orders,  {best_orders['metrics']['releasable_packaging_hours']:>8.1f} hrs,  {best_orders['metrics']['releasable_packaging_qty']:>8} qty
        Malosa Inst:      {best_orders['metrics']['releasable_malosa_instruments_count']:>6} orders,  {best_orders['metrics']['releasable_malosa_instruments_hours']:>8.1f} hrs,  {best_orders['metrics']['releasable_malosa_instruments_qty']:>8} qty
        Total Instruments:{best_orders['metrics']['releasable_instruments_count']:>6} orders,  {best_orders['metrics']['releasable_instruments_hours']:>8.1f} hrs,  {best_orders['metrics']['releasable_instruments_qty']:>8} qty
        
        🎵 VIRTUOSO:
        Virtuoso (3806):   {best_orders['metrics']['releasable_virtuoso_count']:>6} orders,  {best_orders['metrics']['releasable_virtuoso_hours']:>8.1f} hrs,  {best_orders['metrics']['releasable_virtuoso_qty']:>8} qty

   🏆 BEST STRATEGY (Hours): {best_hours['sorting_strategy']}
      → {best_hours['metrics']['releasable_hours']:,.0f}/{best_hours['metrics']['total_hours']:,.0f} hours releasable ({best_hours['metrics']['releasable_hours']/best_hours['metrics']['total_hours']*100:.1f}%)
   
   🏆 BEST STRATEGY (Qty): {best_qty['sorting_strategy']}
      → {best_qty['metrics']['releasable_qty']:,}/{best_qty['metrics']['total_qty']:,} units releasable ({best_qty['metrics']['releasable_qty']/best_qty['metrics']['total_qty']*100:.1f}%)
   
   📉 WORST STRATEGY: {worst_orders['sorting_strategy']}
      → {worst_orders['metrics']['releasable_count']:,} orders releasable
   
   🔺 IMPROVEMENT POTENTIAL: +{improvement_orders:,} more orders ({improvement_pct:.1f}% boost)

"""
            
            summary_text += f"""⏱️ PERFORMANCE METRICS:
   Total Processing Time: {processing_time:.2f} seconds
   Processing Speed: {orders_per_second:.1f} orders/second
   Average per Strategy: {processing_time/len(scenarios_for_comparison):.1f} seconds
   
💾 Results saved to: {os.path.basename(output_file) if output_file else 'No export (Quick Analysis Mode)'}
   
🔥 OPTIMIZATION FEATURES:
   ✓ All sorting strategies tested
   ✓ Triple optimization: Orders + Hours + Quantity
   ✓ Only optimal results saved as individual sheets
   ✓ Complete strategy comparison table
   ✓ Improvement potential analysis"""
            
        elif len(scenarios) > 1:
            # Multi-scenario summary (standard mode)
            best_scenario = max(scenarios, key=lambda s: s['metrics']['releasable_count'])
            worst_scenario = min(scenarios, key=lambda s: s['metrics']['releasable_count'])
            improvement = best_scenario['metrics']['releasable_count'] - worst_scenario['metrics']['releasable_count']
            
            summary_text = f"""✅ MULTI-SCENARIO ANALYSIS COMPLETE!

📊 SCENARIOS COMPARED: {len(scenarios)}

🔧 MATERIAL CATEGORIES PROCESSED:
   Kits (3001, 3801, 5001): {'✓ Included' if include_kits_var.get() else '✗ Excluded'}
   Instruments (3802, 3803, 3804, 3805): {'✓ Included' if include_instruments_var.get() else '✗ Excluded'}
   Virtuoso (3806): {'✓ Included' if include_virtuoso_var.get() else '✗ Excluded'}
   Kit Samples (KIT SAMPLES): {'✓ Included' if include_kit_samples_var.get() else '✗ Excluded'}

🏆 BEST PERFORMER: {os.path.basename(best_scenario['filepath'])}
   ✅ {best_scenario['metrics']['releasable_count']:,} releasable orders ({best_scenario['metrics']['releasable_count']/best_scenario['metrics']['total_orders']*100:.1f}%)
   🔧 BVI Kits: {best_scenario['metrics']['releasable_bvi_kits_count']:,} orders, {best_scenario['metrics']['releasable_bvi_kits_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_bvi_kits_qty']:,} qty
   🔧 Malosa Kits: {best_scenario['metrics']['releasable_malosa_kits_count']:,} orders, {best_scenario['metrics']['releasable_malosa_kits_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_malosa_kits_qty']:,} qty
   🔬 Manufacturing: {best_scenario['metrics']['releasable_manufacturing_count']:,} orders, {best_scenario['metrics']['releasable_manufacturing_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_manufacturing_qty']:,} qty
   🔧 Assembly: {best_scenario['metrics']['releasable_assembly_count']:,} orders, {best_scenario['metrics']['releasable_assembly_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_assembly_qty']:,} qty
   📦 Packaging: {best_scenario['metrics']['releasable_packaging_count']:,} orders, {best_scenario['metrics']['releasable_packaging_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_packaging_qty']:,} qty
   🔬 Malosa Instruments: {best_scenario['metrics']['releasable_malosa_instruments_count']:,} orders, {best_scenario['metrics']['releasable_malosa_instruments_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_malosa_instruments_qty']:,} qty
   📊 Total Instruments: {best_scenario['metrics']['releasable_instruments_count']:,} orders, {best_scenario['metrics']['releasable_instruments_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_instruments_qty']:,} qty

🎵 VIRTUOSO:
   Virtuoso (3806): {best_scenario['metrics']['releasable_virtuoso_count']:,} orders, {best_scenario['metrics']['releasable_virtuoso_hours']:,.0f} hrs, {best_scenario['metrics']['releasable_virtuoso_qty']:,} qty

📉 BASELINE: {os.path.basename(worst_scenario['filepath'])}
   ✅ {worst_scenario['metrics']['releasable_count']:,} releasable orders ({worst_scenario['metrics']['releasable_count']/worst_scenario['metrics']['total_orders']*100:.1f}%)

🔺 IMPROVEMENT: +{improvement:,} more orders releasable

⏱️ PERFORMANCE METRICS:
   Total Processing Time: {processing_time:.2f} seconds
   Processing Speed: {orders_per_second:.1f} orders/second
   
💾 Results saved to: {os.path.basename(output_file) if output_file else 'No export (Quick Analysis Mode)'}"""
        else:
            #  scenario summary
            scenario = scenarios[0]
            metrics = scenario['metrics']
            
            summary_text = f"""✅ PROCESSING COMPLETE!

📊 RESULTS SUMMARY:
   Total Orders:     {format_metric(safe_metric(metrics, 'total_orders')):>8}
   ✅ Releasable:    {format_metric(safe_metric(metrics, 'releasable_count')):>8} ({format_metric(safe_metric(metrics, 'releasable_count') / safe_metric(metrics, 'total_orders') * 100, 'percentage')})
   ❌ On Hold:       {format_metric(safe_metric(metrics, 'held_count')):>8} ({format_metric(safe_metric(metrics, 'held_count') / safe_metric(metrics, 'total_orders') * 100, 'percentage')})
   🏷️ Piggyback:     {format_metric(safe_metric(metrics, 'pb_count')):>8}
   ⚠️ Skipped:       {format_metric(safe_metric(metrics, 'skipped_count')):>8}

🔧 MATERIAL CATEGORIES PROCESSED:
   Kits (3001, 3801, 5001): {'✓ Included' if include_kits_var.get() else '✗ Excluded'}
   Instruments (3802, 3803, 3804, 3805): {'✓ Included' if include_instruments_var.get() else '✗ Excluded'}
   Virtuoso (3806): {'✓ Included' if include_virtuoso_var.get() else '✗ Excluded'}
   Kit Samples (KIT SAMPLES): {'✓ Included' if include_kit_samples_var.get() else '✗ Excluded'}

🔧 RELEASABLE KITS:
   BVI Kits (3001, 3801):    {format_metric(safe_metric(metrics, 'releasable_bvi_kits_count')):>6} orders,  {format_metric(safe_metric(metrics, 'releasable_bvi_kits_hours'), 'hours'):>8} hrs,  {format_metric(safe_metric(metrics, 'releasable_bvi_kits_qty')):>8} qty
   Malosa Kits (5001):       {format_metric(safe_metric(metrics, 'releasable_malosa_kits_count')):>6} orders,  {format_metric(safe_metric(metrics, 'releasable_malosa_kits_hours'), 'hours'):>8} hrs,  {format_metric(safe_metric(metrics, 'releasable_malosa_kits_qty')):>8} qty
   Total Kits:               {format_metric(safe_metric(metrics, 'releasable_kits_count')):>6} orders,  {format_metric(safe_metric(metrics, 'releasable_kits_hours'), 'hours'):>8} hrs,  {format_metric(safe_metric(metrics, 'releasable_kits_qty')):>8} qty

🔬 RELEASABLE INSTRUMENTS:
   Manufacturing (3802):     {format_metric(safe_metric(metrics, 'releasable_manufacturing_count')):>6} orders,  {format_metric(safe_metric(metrics, 'releasable_manufacturing_hours'), 'hours'):>8} hrs,  {format_metric(safe_metric(metrics, 'releasable_manufacturing_qty')):>8} qty
   Assembly (3803):          {format_metric(safe_metric(metrics, 'releasable_assembly_count')):>6} orders,  {format_metric(safe_metric(metrics, 'releasable_assembly_hours'), 'hours'):>8} hrs,  {format_metric(safe_metric(metrics, 'releasable_assembly_qty')):>8} qty
   Packaging (3804):         {format_metric(safe_metric(metrics, 'releasable_packaging_count')):>6} orders,  {format_metric(safe_metric(metrics, 'releasable_packaging_hours'), 'hours'):>8} hrs,  {format_metric(safe_metric(metrics, 'releasable_packaging_qty')):>8} qty
   Malosa Instruments (3805):{format_metric(safe_metric(metrics, 'releasable_malosa_instruments_count')):>6} orders,  {format_metric(safe_metric(metrics, 'releasable_malosa_instruments_hours'), 'hours'):>8} hrs,  {format_metric(safe_metric(metrics, 'releasable_malosa_instruments_qty')):>8} qty
   Total Instruments:        {format_metric(safe_metric(metrics, 'releasable_instruments_count')):>6} orders,  {format_metric(safe_metric(metrics, 'releasable_instruments_hours'), 'hours'):>8} hrs,  {format_metric(safe_metric(metrics, 'releasable_instruments_qty')):>8} qty

🎵 RELEASABLE VIRTUOSO:
   Virtuoso (3806):          {format_metric(safe_metric(metrics, 'releasable_virtuoso_count')):>6} orders,  {format_metric(safe_metric(metrics, 'releasable_virtuoso_hours'), 'hours'):>8} hrs,  {format_metric(safe_metric(metrics, 'releasable_virtuoso_qty')):>8} qty

⏱️ LABOR HOURS SUMMARY:
   Total Hours:              {format_metric(safe_metric(metrics, 'total_hours'), 'hours'):>8}
   ✅ Releasable Hours:       {format_metric(safe_metric(metrics, 'releasable_hours'), 'hours'):>8} ({format_metric(safe_metric(metrics, 'releasable_hours') / safe_metric(metrics, 'total_hours') * 100, 'percentage')})

⏱️ PERFORMANCE METRICS:
   Processing Time: {processing_time:.2f} seconds
   Orders per Second: {orders_per_second:.1f}

💾 Results saved to: {os.path.basename(output_file) if output_file else 'No export (Quick Analysis Mode)'}"""
        
        # Add performance report to the summary
        performance_report = generate_performance_report()
        full_summary = summary_text + "\n\n" + performance_report
        
        results_text.delete(1.0, tk.END)
        results_text.insert(1.0, full_summary)
        
        # For status bar
        if minmax_mode:
            status_var.set(f"🔥 MIN/MAX OPTIMIZATION COMPLETE! {len(scenarios_for_comparison)} strategies tested, {len(scenarios)} best results saved in {processing_time:.1f}s")
        elif len(scenarios) > 1:
            best_scenario = max(scenarios, key=lambda s: s['metrics']['releasable_count'])
            worst_scenario = min(scenarios, key=lambda s: s['metrics']['releasable_count'])
            improvement = best_scenario['metrics']['releasable_count'] - worst_scenario['metrics']['releasable_count']
            status_var.set(f"✅ ALL {len(scenarios)} SCENARIOS COMPLETE! Best: {best_scenario['metrics']['releasable_count']:,} releasable (+{improvement:,} vs worst) | Total time: {processing_time:.1f}s")
        else:
            total_orders = scenarios[0]['metrics']['total_orders']
            total_releasable = scenarios[0]['metrics']['releasable_count']
            status_var.set(f"✅ PROCESSING COMPLETE! {total_releasable:,}/{total_orders:,} orders releasable in {processing_time:.1f}s")
            
        # Set frame color based on mode - blue for quick mode, green for normal mode
        if no_export_var.get():
            main_frame.configure(style='Quick.TFrame')  # Blue for quick mode
        else:
            main_frame.configure(style='Success.TFrame')  # Green for normal mode
        
        # Quick Analysis: Write to BytesIO buffer, don't save to disk
        quick_analysis_excel_buffer = BytesIO()
        with pd.ExcelWriter(quick_analysis_excel_buffer, engine='openpyxl') as writer:
            # ... existing code for writing Excel ...
            # (copy all Excel writing code here, but use writer as above)
            # Define styles once at the start
            header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            separator_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            bold_font = Font(bold=True)
            for scenario in scenarios:
                sheet_name = scenario['name'][:31]
                df = scenario['results_df'].copy()
                numeric_columns = ['Demand', 'Hours']
                for col in numeric_columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(df.columns, 1):
                    col_letter = get_column_letter(idx)
                    if col == 'Hours':
                        for cell in worksheet[col_letter][1:]:
                            cell.number_format = '#,##0.0'
                    elif col == 'Demand':
                        for cell in worksheet[col_letter][1:]:
                            cell.number_format = '#,##0'
            summary_data.to_excel(writer, sheet_name='Summary', index=False)
            worksheet = writer.sheets['Summary']
            for row in range(2, worksheet.max_row + 1):
                value_cell = worksheet.cell(row=row, column=2)
                metric_cell = worksheet.cell(row=row, column=1)
                if any(term in metric_cell.value for term in ['Hours', 'Time']):
                    value_cell.number_format = '#,##0.0'
                elif any(term in metric_cell.value for term in ['Orders', 'Count', 'Quantity']):
                    value_cell.number_format = '#,##0'
                elif 'Rate' in metric_cell.value or 'Speed' in metric_cell.value:
                    value_cell.number_format = '#,##0.0'
                if metric_cell.value.startswith('---'):
                    for col in range(1, 3):
                        cell = worksheet.cell(row=row, column=col)
                        cell.fill = separator_fill
                elif any(metric_cell.value.startswith(prefix) for prefix in ['Total', 'Releasable', 'BVI', 'Malosa', 'Manufacturing', 'Assembly', 'Packaging', 'Virtuoso']):
                    metric_cell.font = bold_font
            for col in range(1, 3):
                cell = worksheet.cell(row=1, column=col)
                cell.font = bold_font
                cell.fill = header_fill
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            if len(scenarios_for_comparison) > 1:
                comparison_df_formatted = comparison_df.copy()
                numeric_columns = [
                    'Total Orders', 'Releasable Orders', 'Held Orders', 'Piggyback Orders',
                    'Total Hours', 'Releasable Hours', 'BVI Kits', 'BVI Kit Hours', 'BVI Kit Qty',
                    'Malosa Kits', 'Malosa Kit Hours', 'Malosa Kit Qty', 'Total Kits', 'Total Kit Hours',
                    'Total Kit Qty', 'Manufacturing (3802)', 'Manufacturing Hours', 'Manufacturing Qty',
                    'Assembly (3803)', 'Assembly Hours', 'Assembly Qty', 'Packaging (3804)', 'Packaging Hours',
                    'Packaging Qty', 'Malosa Inst (3805)', 'Malosa Inst Hours', 'Malosa Inst Qty',
                    'Virtuoso (3806)', 'Virtuoso Hours', 'Virtuoso Qty', 'Total Instruments',
                    'Total Inst Hours', 'Total Inst Qty', 'Committed Parts', 'Committed Qty',
                    'Kit Samples', 'Kit Samples Hours', 'Kit Samples Qty'
                ]
                for col in numeric_columns:
                    if col in comparison_df_formatted.columns:
                        comparison_df_formatted[col] = comparison_df_formatted[col].astype(str).str.replace(',', '').str.replace('$', '')
                        comparison_df_formatted[col] = pd.to_numeric(comparison_df_formatted[col], errors='coerce')
                pct_columns = ['Release Rate (%)', 'Qty Release Rate (%)', 'Labor Release Rate (%)']
                for col in pct_columns:
                    if col in comparison_df_formatted.columns:
                        comparison_df_formatted[col] = comparison_df_formatted[col].astype(str).str.rstrip('%').astype(float) / 100
                comparison_df_formatted.to_excel(writer, sheet_name='Strategy Comparison', index=False)
                comp_worksheet = writer.sheets['Strategy Comparison']
                for col_idx, col_name in enumerate(comparison_df_formatted.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    if col_name in pct_columns:
                        for cell in comp_worksheet[col_letter][1:]:
                            cell.number_format = '0.0%'
                    elif 'Hours' in str(col_name):
                        for cell in comp_worksheet[col_letter][1:]:
                            cell.number_format = '#,##0.0'
                    elif any(term in str(col_name) for term in ['Orders', 'Qty', 'Count', 'Parts']):
                        for cell in comp_worksheet[col_letter][1:]:
                            cell.number_format = '#,##0'
                for col in range(1, comp_worksheet.max_column + 1):
                    cell = comp_worksheet.cell(row=1, column=col)
                    cell.font = bold_font
                    cell.fill = header_fill
                    max_length = 0
                    column = [cell for cell in comp_worksheet[get_column_letter(col)]]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                        adjusted_width = min((max_length + 2), 50)
                        comp_worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width
            quick_analysis_excel_buffer.seek(0)
            output_file = None
        # ... existing code ...
        # Show/hide download button
        if no_export_var.get():
            download_btn.grid()
        else:
            download_btn.grid_remove()
        # ... existing code ...
    except Exception as e:
        messagebox.showerror("Error", f"Processing failed:\n\n{str(e)}")
        status_var.set("❌ Processing failed")
        download_btn.grid_remove()

def generate_performance_report():
    """Generate detailed performance report with phase breakdown"""
    phase_summary = performance_tracker.get_phase_summary()
    memory_summary = performance_tracker.get_memory_summary()
    
    report = []
    report.append("🔍 DETAILED PERFORMANCE ANALYSIS")
    report.append("=" * 50)
    
    # Phase breakdown
    report.append("\n📊 PHASE BREAKDOWN:")
    total_time = 0
    for phase, metrics in phase_summary.items():
        total_time += metrics['total_time']
        report.append(f"   {phase}:")
        report.append(f"     Total Time: {metrics['total_time']:.3f}s")
        report.append(f"     Average Time: {metrics['avg_time']:.3f}s")
        report.append(f"     Count: {metrics['count']}")
        report.append(f"     Min/Max: {metrics['min_time']:.3f}s / {metrics['max_time']:.3f}s")
    
    # Calculate percentages
    report.append(f"\n📈 TIME DISTRIBUTION:")
    for phase, metrics in phase_summary.items():
        percentage = (metrics['total_time'] / total_time * 100) if total_time > 0 else 0
        report.append(f"   {phase}: {percentage:.1f}% ({metrics['total_time']:.3f}s)")
    
    # Memory usage
    if memory_summary:
        report.append(f"\n💾 MEMORY USAGE:")
        report.append(f"   Peak Memory: {memory_summary['peak_memory_mb']:.1f} MB")
        report.append(f"   Average Memory: {memory_summary['avg_memory_mb']:.1f} MB")
        report.append(f"   Initial Memory: {memory_summary['initial_memory_mb']:.1f} MB")
        report.append(f"   Final Memory: {memory_summary['final_memory_mb']:.1f} MB")
        report.append(f"   Memory Growth: {memory_summary['final_memory_mb'] - memory_summary['initial_memory_mb']:.1f} MB")
    
    # Performance insights
    report.append(f"\n💡 PERFORMANCE INSIGHTS:")
    if phase_summary:
        slowest_phase = max(phase_summary.items(), key=lambda x: x[1]['total_time'])
        fastest_phase = min(phase_summary.items(), key=lambda x: x[1]['total_time'])
        
        report.append(f"   Slowest Phase: {slowest_phase[0]} ({slowest_phase[1]['total_time']:.3f}s)")
        report.append(f"   Fastest Phase: {fastest_phase[0]} ({fastest_phase[1]['total_time']:.3f}s)")
        
        # Bottleneck analysis
        if slowest_phase[1]['total_time'] > total_time * 0.5:  # If slowest phase is >50% of total time
            report.append(f"   ⚠️  BOTTLENECK DETECTED: {slowest_phase[0]} is consuming {slowest_phase[1]['total_time']/total_time*100:.1f}% of total time")
        
        # Database vs Processing analysis with improved categorization
        db_phases = [p for p in phase_summary.keys() if any(term in p for term in ['Load', 'Database', 'Connection'])]
        processing_phases = [p for p in phase_summary.keys() if p not in db_phases and p != 'Total Processing']
        
        db_time = sum(phase_summary[p]['total_time'] for p in db_phases)
        processing_time = sum(phase_summary[p]['total_time'] for p in processing_phases)
        
        report.append(f"\n   Database Operations: {db_time:.3f}s ({db_time/total_time*100:.1f}%)")
        report.append(f"   Processing Operations: {processing_time:.3f}s ({processing_time/total_time*100:.1f}%)")
        
        # Additional insights for database operations
        if db_phases:
            report.append(f"\n   📊 DATABASE BREAKDOWN:")
            for phase in db_phases:
                phase_time = phase_summary[phase]['total_time']
                phase_pct = (phase_time / db_time * 100) if db_time > 0 else 0
                report.append(f"     {phase}: {phase_time:.3f}s ({phase_pct:.1f}% of DB time)")
        
        # Additional insights for processing operations
        if processing_phases:
            report.append(f"\n   🔧 PROCESSING BREAKDOWN:")
            for phase in processing_phases:
                phase_time = phase_summary[phase]['total_time']
                phase_pct = (phase_time / processing_time * 100) if processing_time > 0 else 0
                report.append(f"     {phase}: {phase_time:.3f}s ({phase_pct:.1f}% of processing time)")
    
    return "\n".join(report)

def copy_summary_to_clipboard():
    summary_text_content = results_text.get("1.0", tk.END).strip()
    root.clipboard_clear()
    root.clipboard_append(summary_text_content)
    root.update()
    copy_btn.config(text="✅ Copied!")
    root.after(2000, lambda: copy_btn.config(text="📋 Copy Summary"))

# Create GUI
root = tk.Tk()
root.title(f"PlanSnap {VERSION} - Material Release Planning Tool")
root.geometry("600x650")

# Global buffer for quick analysis Excel file
quick_analysis_excel_buffer = None

# Main frame
main_frame = ttk.Frame(root, padding="20")
main_frame.grid(row=0, column=0, sticky="nsew")

# Title
title_label = ttk.Label(main_frame, text=f"PlanSnap {VERSION}", 
                       font=('Arial', 16, 'bold'))
title_label.grid(row=0, column=0, pady=(0, 5))

# Version date
version_label = ttk.Label(main_frame, text=f"Updated: {VERSION_DATE}", 
                         font=('Arial', 8))
version_label.grid(row=1, column=0, pady=(0, 15))

# Instructions
#instructions = """• Database connection configured via db_credentials.env
#• Enable Min/Max Optimization to find the best sorting strategy for database data
#• Use Quick Analysis Mode for faster results without Excel export
#• Select specific material categories to process (Kits, Instruments, Virtuoso)"""

#inst_label = ttk.Label(main_frame, text=instructions, justify=tk.LEFT, wraplength=700)
#inst_label.grid(row=2, column=0, pady=(0, 20))

# Min/Max Mode checkbox with tooltip
minmax_frame = ttk.Frame(main_frame)
minmax_frame.grid(row=3, column=0, pady=(0, 10))

minmax_var = tk.BooleanVar()
minmax_checkbox = ttk.Checkbutton(
    minmax_frame, 
    text="🔥 Enable Triple Optimization Mode",
    variable=minmax_var,
    style='Big.TCheckbutton'
)
minmax_checkbox.grid(row=0, column=0)

minmax_tooltip = ttk.Label(
    minmax_frame,
    text="Tests all sorting strategies to find best results for Orders, Hours, and Quantity",
    font=('Arial', 8, 'italic'),
    foreground='gray'
)
minmax_tooltip.grid(row=1, column=0, pady=(0, 10))

# No Export checkbox with tooltip
no_export_frame = ttk.Frame(main_frame)
no_export_frame.grid(row=4, column=0, pady=(0, 10))

no_export_var = tk.BooleanVar()
no_export_checkbox = ttk.Checkbutton(
    no_export_frame, 
    text="⚡ Quick Analysis Mode",
    variable=no_export_var,
    style='Big.TCheckbutton'
)
no_export_checkbox.grid(row=0, column=0)

no_export_tooltip = ttk.Label(
    no_export_frame,
    text="Show results instantly without creating Excel files (useful for rapid testing)",
    font=('Arial', 8, 'italic'),
    foreground='gray'
)
no_export_tooltip.grid(row=1, column=0, pady=(0, 10))

# Material Category Selection
material_frame = ttk.LabelFrame(main_frame, text="🔧 Material Categories to Process", padding="10")
material_frame.grid(row=5, column=0, pady=(0, 10), sticky="ew")

# Create variables for material category checkboxes
include_kits_var = tk.BooleanVar(value=True)
include_instruments_var = tk.BooleanVar(value=True)
include_virtuoso_var = tk.BooleanVar(value=True)
include_kit_samples_var = tk.BooleanVar(value=False)

# Kits checkbox
kits_checkbox = ttk.Checkbutton(
    material_frame,
    text="🔧 Kits (Planner codes: 3001, 3801, 5001)",
    variable=include_kits_var,
    style='Big.TCheckbutton'
)
kits_checkbox.grid(row=0, column=0, sticky=tk.W, pady=(0, 5))

# Instruments checkbox
instruments_checkbox = ttk.Checkbutton(
    material_frame,
    text="🔬 Instruments (Planner codes: 3802, 3803, 3804, 3805)",
    variable=include_instruments_var,
    style='Big.TCheckbutton'
)
instruments_checkbox.grid(row=1, column=0, sticky=tk.W, pady=(0, 5))

# Virtuoso checkbox
virtuoso_checkbox = ttk.Checkbutton(
    material_frame,
    text="🎵 Virtuoso (Planner code: 3806)",
    variable=include_virtuoso_var,
    style='Big.TCheckbutton'
)
virtuoso_checkbox.grid(row=2, column=0, sticky=tk.W, pady=(0, 5))

# Kit Samples checkbox
kit_samples_checkbox = ttk.Checkbutton(
    material_frame,
    text="🧪 Kit Samples (Planner code: KIT SAMPLES)",
    variable=include_kit_samples_var,
    style='Big.TCheckbutton'
)
kit_samples_checkbox.grid(row=3, column=0, sticky=tk.W, pady=(0, 5))

# Material categories tooltip
material_tooltip = ttk.Label(
    material_frame,
    text="Untick categories to exclude them from material availability checks",
    font=('Arial', 8, 'italic'),
    foreground='gray'
)
material_tooltip.grid(row=4, column=0, pady=(5, 0), sticky=tk.W)

# Process button
process_btn = ttk.Button(main_frame, text="🗄️ CONNECT TO DATABASE & PROCESS", 
                        command=load_and_process_database, 
                        style='Big.TButton')
process_btn.grid(row=6, column=0, pady=(10, 20))

# Configure button style
style = ttk.Style()
style.configure('Big.TButton', font=('Arial', 12, 'bold'))
style.configure('Big.TCheckbutton', font=('Arial', 10, 'bold'))
style.configure('Success.TFrame', background='#7ff09a')  # Green for normal mode
style.configure('Quick.TFrame', background='#87ceeb')    # Blue for quick mode

# Status
status_var = tk.StringVar()
status_var.set("🔄 Ready - Connect to database to begin processing")
status_label = ttk.Label(main_frame, textvariable=status_var, font=('Arial', 10))
status_label.grid(row=7, column=0, pady=(0, 10), sticky=tk.W)

# Results area
results_frame = ttk.LabelFrame(main_frame, text="📊 Results", padding="10")
results_frame.grid(row=8, column=0, sticky="nsew", pady=(10, 0))

results_text = tk.Text(results_frame, height=18, width=90, font=('Consolas', 9))
scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=results_text.yview)

results_text.configure(yscrollcommand=scrollbar.set)

results_text.grid(row=0, column=0, sticky="nsew")
scrollbar.grid(row=0, column=1, sticky="ns")

copy_btn = ttk.Button(main_frame, text="📋 Copy Summary", command=copy_summary_to_clipboard)
copy_btn.grid(row=9, column=0, pady=(10, 10))

# Download button for Quick Analysis
def download_quick_analysis_file():
    global quick_analysis_excel_buffer
    if quick_analysis_excel_buffer is None:
        messagebox.showerror("No file", "No quick analysis file available.")
        return
    
    try:
        # Check if buffer has data
        if quick_analysis_excel_buffer.getvalue() == b'':
            messagebox.showerror("No data", "Quick analysis buffer is empty.")
            return
            
        # Try to get a default filename
        default_filename = f"Quick_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Quick Analysis Results",
            initialfile=default_filename
        )
        
        if file_path:
            # Ensure the buffer is at the beginning
            quick_analysis_excel_buffer.seek(0)
            
            # Get the buffer data
            buffer_data = quick_analysis_excel_buffer.getvalue()
            
            # Check if we have data to write
            if not buffer_data:
                messagebox.showerror("Error", "No data available to save.")
                return
            
            # Write the file with proper error handling
            with open(file_path, "wb") as f:
                f.write(buffer_data)
            
        else:
            # User cancelled the save dialog
            pass
            
    except Exception as e:
        # Try fallback approach for executable
        try:
            # Try to save to desktop as fallback
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            if not os.path.exists(desktop_path):
                desktop_path = os.path.join(os.path.expanduser("~"), "OneDrive - BVI\\Desktop")
            
            fallback_filename = f"Quick_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            fallback_path = os.path.join(desktop_path, fallback_filename)
            
            quick_analysis_excel_buffer.seek(0)
            buffer_data = quick_analysis_excel_buffer.getvalue()
            
            with open(fallback_path, "wb") as f:
                f.write(buffer_data)
            
            messagebox.showinfo("Success", f"File saved to desktop:\n{fallback_path}")
            
        except Exception as fallback_error:
            messagebox.showerror("Error", f"Failed to save file:\n{str(e)}\n\nFallback also failed:\n{str(fallback_error)}")
            # For debugging in executable
            print(f"Error saving file: {str(e)}")
            print(f"Fallback error: {str(fallback_error)}")

# Add the download button, initially hidden
download_btn = ttk.Button(main_frame, text="⬇️ Download File", command=download_quick_analysis_file)
download_btn.grid(row=10, column=0, pady=(10, 10))
download_btn.grid_remove()

# Configure grid weights
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
main_frame.columnconfigure(0, weight=1)
main_frame.rowconfigure(8, weight=1)
results_frame.columnconfigure(0, weight=1)
results_frame.rowconfigure(0, weight=1)

# Show initial message
results_text.insert(1.0, "Connect to your database to begin material release planning...")

if __name__ == "__main__":
    root.mainloop()
